from flask import Flask, render_template, request, redirect, url_for, session, send_file, make_response
from flask_cors import CORS
from fpdf import FPDF
import gspread
from oauth2client.service_account import ServiceAccountCredentials
from datetime import datetime, timedelta
import pandas as pd
import io
import os
import gspread.utils 
import json
import time
import calendar
import requests 
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
from openpyxl.utils import get_column_letter

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'lmt_driver_app_secret_key_2024')
CORS(app)

SPREADSHEET_ID = '15kJuEyhIaIjxZsqvPIxhOqzTrB1eY62KDHRjNITcRkM'  # <--- ‡∏ô‡∏≥ ID ‡∏°‡∏≤‡πÉ‡∏™‡πà‡πÉ‡∏ô‡πÄ‡∏Ñ‡∏£‡∏∑‡πà‡∏≠‡∏á‡∏´‡∏°‡∏≤‡∏¢‡∏Ñ‡∏≥‡∏û‡∏π‡∏î‡∏ô‡∏µ‡πâ ‡πÄ‡∏ä‡πà‡∏ô '1AbCd-EfGhIj...'

# ==========================================
# [CONFIG] ‡∏ï‡∏±‡πâ‡∏á‡∏Ñ‡πà‡∏≤ Discord Webhook
# ==========================================
DISCORD_WEBHOOK_URL = 'https://discord.com/api/webhooks/1444236316404482139/UJc-I_NRT33p9UKCas5ATGgjAlqlrtxBuPhvKYKnI-Pz2_AyxAnOs_UFNl203_sqLsI5'

def send_discord_msg(message):
    """‡∏ü‡∏±‡∏á‡∏Å‡πå‡∏ä‡∏±‡∏ô‡∏™‡πà‡∏á‡∏Ç‡πâ‡∏≠‡∏Ñ‡∏ß‡∏≤‡∏°‡πÄ‡∏Ç‡πâ‡∏≤ Discord"""
    try:
        if not DISCORD_WEBHOOK_URL or '‡∏ß‡∏≤‡∏á_' in DISCORD_WEBHOOK_URL:
            return

        payload = {
            "content": message,
            "username": "LMT Smart Bot",
            "avatar_url": "https://cdn-icons-png.flaticon.com/512/2936/2936956.png"
        }
        requests.post(DISCORD_WEBHOOK_URL, json=payload)
    except Exception as e:
        print(f"Discord Notify Error: {e}")

# --- Caching System ---
cache_storage = {
    'Jobs': {'data': None, 'timestamp': 0},
    'Drivers': {'data': None, 'timestamp': 0},
    'Users': {'data': None, 'timestamp': 0}
}
CACHE_DURATION = 60 

def get_cached_records(sheet, worksheet_name):
    current_time = time.time()
    cache_entry = cache_storage.get(worksheet_name)
    
    if cache_entry and cache_entry['data'] is not None:
        if current_time - cache_entry['timestamp'] < CACHE_DURATION:
            return cache_entry['data']
    
    try:
        data = sheet.worksheet(worksheet_name).get_all_records()
        cache_storage[worksheet_name] = {
            'data': data,
            'timestamp': current_time
        }
        return data
    except gspread.exceptions.APIError as e:
        if "429" in str(e) and cache_entry and cache_entry['data'] is not None:
            return cache_entry['data']
        raise e

def invalidate_cache(worksheet_name):
    if worksheet_name in cache_storage:
        cache_storage[worksheet_name] = {'data': None, 'timestamp': 0}

# --- Helper Functions ---

def get_shift_info(round_time):
    """‡∏£‡∏∞‡∏ö‡∏∏‡∏Å‡∏∞ Day/Night ‡∏à‡∏≤‡∏Å‡πÄ‡∏ß‡∏•‡∏≤"""
    is_day = True
    try:
        h = int(str(round_time).split(':')[0])
        if h < 6 or h >= 19: is_day = False
    except: pass
    return is_day, ("‡∏Å‡∏•‡∏≤‡∏á‡∏ß‡∏±‡∏ô ‚òÄÔ∏è" if is_day else "‡∏Å‡∏•‡∏≤‡∏á‡∏Ñ‡∏∑‡∏ô üåô")

def get_driver_details(sheet, driver_name):
    """‡∏î‡∏∂‡∏á‡πÄ‡∏•‡∏Ç‡∏ö‡∏±‡∏ï‡∏£‡πÅ‡∏•‡∏∞‡πÄ‡∏ö‡∏≠‡∏£‡πå‡πÇ‡∏ó‡∏£‡∏à‡∏≤‡∏Å Cache Drivers"""
    try:
        drivers = get_cached_records(sheet, 'Drivers')
        for d in drivers:
            if d.get('Name') == driver_name:
                return d.get('ID_Card', '-'), d.get('Phone', '-')
    except: pass
    return '-', '-'
    
def comma_format(value):
    """‡πÅ‡∏õ‡∏•‡∏á‡∏ï‡∏±‡∏ß‡πÄ‡∏•‡∏Ç‡πÉ‡∏´‡πâ‡∏°‡∏µ‡πÄ‡∏Ñ‡∏£‡∏∑‡πà‡∏≠‡∏á‡∏´‡∏°‡∏≤‡∏¢ ,"""
    if not value: return ""
    if value == '-': return "-"
    try:
        clean_val = str(value).replace(',', '')
        num = float(clean_val)
        if num.is_integer():
            return "{:,.0f}".format(num)
        return "{:,.2f}".format(num)
    except:
        return str(value)

def thai_date_filter(date_val):
    """‡πÅ‡∏õ‡∏•‡∏á‡∏ß‡∏±‡∏ô‡∏ó‡∏µ‡πà string (YYYY-MM-DD) ‡πÄ‡∏õ‡πá‡∏ô‡∏£‡∏π‡∏õ‡πÅ‡∏ö‡∏ö‡πÑ‡∏ó‡∏¢‡∏ï‡∏±‡∏ß‡πÄ‡∏•‡∏Ç (30/11/2568)"""
    if not date_val: return ""
    try:
        if isinstance(date_val, str):
            d = datetime.strptime(str(date_val).strip(), "%Y-%m-%d")
        else:
            d = date_val
            
        year = d.year + 543
        return f"{d.day:02d}/{d.month:02d}/{year}"
    except:
        return str(date_val)
        
def parse_po_data(po_str, doc_str, weight_str):
    """
    ‡πÅ‡∏õ‡∏•‡∏á String ‡∏à‡∏≤‡∏Å Database ‡πÉ‡∏´‡πâ‡πÄ‡∏õ‡πá‡∏ô List of Dict ‡πÄ‡∏û‡∏∑‡πà‡∏≠‡πÅ‡∏™‡∏î‡∏á‡∏ú‡∏•
    Input: "PO1,PO2", "PO1:Doc1 | PO2:Doc2", "PO1:10 | PO2:20"
    """
    if not po_str: return []
    
    # ‡πÅ‡∏õ‡∏•‡∏á PO String ‡πÄ‡∏õ‡πá‡∏ô List
    po_list = [p.strip() for p in po_str.split(',') if p.strip()]
    
    # ‡πÅ‡∏õ‡∏•‡∏á Doc ‡πÅ‡∏•‡∏∞ Weight ‡πÄ‡∏õ‡πá‡∏ô Dictionary
    doc_map = {}
    if doc_str:
        parts = doc_str.split('|')
        for p in parts:
            if ':' in p:
                k, v = p.split(':', 1)
                doc_map[k.strip()] = v.strip()

    weight_map = {}
    if weight_str:
        parts = weight_str.split('|')
        for p in parts:
            if ':' in p:
                k, v = p.split(':', 1)
                weight_map[k.strip()] = v.strip()

    # ‡∏£‡∏ß‡∏°‡∏£‡πà‡∏≤‡∏á
    result = []
    for po in po_list:
        result.append({
            'name': po,
            'doc': doc_map.get(po, ''),
            'weight': weight_map.get(po, '')
        })
    return result

# Register Filters
app.jinja_env.filters['comma_format'] = comma_format
app.jinja_env.filters['thai_date'] = thai_date_filter

# ==========================================
# [FIX for Vercel] ‡∏£‡∏∞‡∏ö‡∏ö‡∏à‡∏≥‡∏Ñ‡πà‡∏≤‡∏ú‡πà‡∏≤‡∏ô Google Sheet
# ==========================================
def is_already_notified(sheet, key):
    try:
        try:
            ws_log = sheet.worksheet('NotifyLogs')
        except:
            ws_log = sheet.add_worksheet(title="NotifyLogs", rows=1000, cols=2)
            ws_log.append_row(['Notify_Key', 'Timestamp'])
            return False

        existing_keys = ws_log.col_values(1)
        
        if key in existing_keys:
            return True
        else:
            timestamp = (datetime.now() + timedelta(hours=7)).strftime("%Y-%m-%d %H:%M:%S")
            ws_log.append_row([key, timestamp])
            return False
            
    except Exception as e:
        print(f"Log Sheet Error: {e}")
        return True

# --- Notification Logic ---
def notify_individual_movement(sheet, job_data, step):
    try:
        id_card, phone = get_driver_details(sheet, job_data['Driver'])
        is_day, shift_name = get_shift_info(job_data['Round'])
        
        action_txt = ""
        icon = ""
        if step == '1':
            action_txt = "‡πÄ‡∏Ç‡πâ‡∏≤‡∏ñ‡∏∂‡∏á‡πÇ‡∏£‡∏á‡∏á‡∏≤‡∏ô"
            icon = "üü©" 
        elif step == '6':
            action_txt = "‡∏≠‡∏≠‡∏Å‡∏à‡∏≤‡∏Å‡πÇ‡∏£‡∏á‡∏á‡∏≤‡∏ô"
            icon = "üöÄ" 
        else:
            return

        msg = (
            f"{icon} **‡πÅ‡∏à‡πâ‡∏á‡πÄ‡∏ï‡∏∑‡∏≠‡∏ô: ‡∏£‡∏ñ{action_txt}**\n"
            f"üìÖ PO: `{job_data['PO_Date']}` | ‡∏Å‡∏∞: {shift_name}\n"
            f"‚è∞ ‡∏£‡∏≠‡∏ö‡πÇ‡∏´‡∏•‡∏î: `{job_data['Round']}` | ‡∏Ñ‡∏±‡∏ô‡∏ó‡∏µ‡πà: `{job_data['Car_No']}`\n"
            f"üöõ ‡∏ó‡∏∞‡πÄ‡∏ö‡∏µ‡∏¢‡∏ô: `{job_data['Plate']}`\n"
            f"----------------------------------\n"
            f"üë§ ‡∏Ñ‡∏ô‡∏Ç‡∏±‡∏ö: **{job_data['Driver']}**\n"
            f"üÜî ‡∏ö‡∏±‡∏ï‡∏£: `{id_card}`\n"
            f"üìû ‡πÇ‡∏ó‡∏£: `{phone}`"
        )
        send_discord_msg(msg)
    except Exception as e:
        print(f"Individual Notify Error: {e}")
        
def notify_car_completion(sheet, job_data):
    try:
        all_jobs = sheet.worksheet('Jobs').get_all_records()
        my_trip = [
            j for j in all_jobs
            if str(j['PO_Date']) == str(job_data['PO_Date'])
            and str(j['Round']) == str(job_data['Round'])
            and str(j['Car_No']) == str(job_data['Car_No'])
            and str(j.get('Status', '')).lower() != 'cancel'
        ]
        
        if not my_trip: return
        if not all(j['Status'] == 'Done' for j in my_trip): return

        id_card, phone = get_driver_details(sheet, job_data['Driver'])
        is_day, shift_name = get_shift_info(job_data['Round'])
        
        msg = (
            f"üèÅ **‡πÅ‡∏à‡πâ‡∏á‡πÄ‡∏ï‡∏∑‡∏≠‡∏ô: ‡∏£‡∏ñ‡∏à‡∏ö‡∏á‡∏≤‡∏ô‡∏Ñ‡∏£‡∏ö‡∏ó‡∏∏‡∏Å‡∏™‡∏≤‡∏Ç‡∏≤**\n"
            f"üìÖ PO: `{job_data['PO_Date']}` | ‡∏Å‡∏∞: {shift_name}\n"
            f"‚è∞ ‡∏£‡∏≠‡∏ö‡πÇ‡∏´‡∏•‡∏î: `{job_data['Round']}` | ‡∏Ñ‡∏±‡∏ô‡∏ó‡∏µ‡πà: `{job_data['Car_No']}`\n"
            f"üöõ ‡∏ó‡∏∞‡πÄ‡∏ö‡∏µ‡∏¢‡∏ô: `{job_data['Plate']}`\n"
            f"----------------------------------\n"
            f"üë§ ‡∏Ñ‡∏ô‡∏Ç‡∏±‡∏ö: **{job_data['Driver']}**\n"
            f"üÜî ‡∏ö‡∏±‡∏ï‡∏£: `{id_card}`\n"
            f"üìû ‡πÇ‡∏ó‡∏£: `{phone}`"
        )
        send_discord_msg(msg)
    except Exception as e:
        print(f"Car Completion Notify Error: {e}")

def check_group_completion(sheet, target_po_date, target_round_time, trigger_step):
    try:
        target_is_day, shift_name = get_shift_info(target_round_time)
        raw_jobs = sheet.worksheet('Jobs').get_all_records()
        target_jobs = [j for j in raw_jobs if str(j['PO_Date']).strip() == str(target_po_date).strip()]

        stats = {'total': 0, 'in': 0, 'out': 0, 'done': 0}
        trips = {}
        for job in target_jobs:
            if str(job.get('Status', '')).lower() == 'cancel': continue
            key = (str(job['PO_Date']), str(job['Round']), str(job['Car_No']))
            if key not in trips: trips[key] = []
            trips[key].append(job)

        for key, job_list in trips.items():
            first_job = job_list[0]
            r_time = str(first_job.get('Round', '')).strip()
            is_day, _ = get_shift_info(r_time)
            
            if is_day == target_is_day:
                stats['total'] += 1
                if str(first_job.get('T1_Enter', '')).strip() != '': stats['in'] += 1
                if str(first_job.get('T6_Exit', '')).strip() != '': stats['out'] += 1
                if all(j['Status'] == 'Done' for j in job_list): stats['done'] += 1

        if stats['total'] == 0: return

        now_str = (datetime.now() + timedelta(hours=7)).strftime('%H:%M')
        base_msg = f"‚úÖ PO: {target_po_date} ({shift_name})\nüöõ ‡∏à‡∏≥‡∏ô‡∏ß‡∏ô‡∏ó‡∏±‡πâ‡∏á‡∏´‡∏°‡∏î: `{stats['total']}` ‡∏Ñ‡∏±‡∏ô\nüïí ‡πÄ‡∏ß‡∏•‡∏≤: `{now_str} ‡∏ô.`"
        shift_key = 'day' if target_is_day else 'night'

        if trigger_step == '1':
            if stats['total'] == stats['in'] and stats['total'] > 0:
                cache_key = f"completed_in_{target_po_date}_{shift_key}"
                if not is_already_notified(sheet, cache_key):
                    send_discord_msg(f"üèÅ **‡∏™‡∏£‡∏∏‡∏õ: ‡∏£‡∏ñ‡πÄ‡∏Ç‡πâ‡∏≤‡πÇ‡∏£‡∏á‡∏á‡∏≤‡∏ô ‡∏Ñ‡∏£‡∏ö‡πÅ‡∏•‡πâ‡∏ß!**\n{base_msg}")

        if trigger_step == '6':
            if stats['total'] == stats['out'] and stats['total'] > 0:
                cache_key = f"completed_out_{target_po_date}_{shift_key}"
                if not is_already_notified(sheet, cache_key):
                    send_discord_msg(f"üõ´ **‡∏™‡∏£‡∏∏‡∏õ: ‡∏£‡∏ñ‡∏≠‡∏≠‡∏Å‡∏à‡∏≤‡∏Å‡πÇ‡∏£‡∏á‡∏á‡∏≤‡∏ô ‡∏Ñ‡∏£‡∏ö‡πÅ‡∏•‡πâ‡∏ß!**\n{base_msg}")

        if trigger_step == '8':
            if stats['total'] == stats['done'] and stats['total'] > 0:
                cache_key = f"completed_done_{target_po_date}_{shift_key}"
                if not is_already_notified(sheet, cache_key):
                    send_discord_msg(f"üéâ **‡∏™‡∏£‡∏∏‡∏õ: ‡∏à‡∏ö‡∏á‡∏≤‡∏ô‡∏™‡πà‡∏á‡∏Ç‡∏≠‡∏á ‡∏Ñ‡∏£‡∏ö‡∏ó‡∏∏‡∏Å‡∏Ñ‡∏±‡∏ô‡πÅ‡∏•‡πâ‡∏ß!**\n{base_msg}")

    except Exception as e:
        print(f"Group Notify Error: {e}")

def check_late_and_notify(sheet):
    try:
        now_thai = datetime.now() + timedelta(hours=7)
        raw_jobs = get_cached_records(sheet, 'Jobs')
        unique_cars = {}
        
        active_jobs = [
            j for j in raw_jobs 
            if str(j.get('Status', '')).lower() != 'cancel' 
            and str(j.get('Status', '')).lower() != 'done' 
            and str(j.get('T1_Enter', '')).strip() == ''
        ]

        late_list = {'day': [], 'night': []}

        for job in active_jobs:
            try:
                key = (str(job['PO_Date']), str(job['Round']), str(job['Car_No']))
                if key in unique_cars: continue
                unique_cars[key] = True

                load_date_str = job.get('Load_Date', job['PO_Date'])
                round_str = str(job['Round']).strip()
                
                try:
                    plan_dt = datetime.strptime(f"{load_date_str} {round_str}", "%Y-%m-%d %H:%M")
                except ValueError:
                    continue

                if str(job.get('Load_Date', '')).strip() == '' or str(job.get('Load_Date')) == str(job['PO_Date']):
                    h_plan = plan_dt.hour
                    if 0 <= h_plan < 6:
                        plan_dt = plan_dt + timedelta(days=1)

                if now_thai > plan_dt and (now_thai - plan_dt).total_seconds() < 48 * 3600:
                    diff = now_thai - plan_dt
                    hours_late = diff.total_seconds() / 3600
                    
                    if hours_late >= 2:
                        is_day, _ = get_shift_info(round_str)
                        id_card, phone = get_driver_details(sheet, job['Driver'])
                        minutes_late = int((diff.total_seconds() % 3600) // 60)
                        
                        info_txt = (f"‚Ä¢ ‡∏Ñ‡∏±‡∏ô‡∏ó‡∏µ‡πà {job['Car_No']} (‡∏ô‡∏±‡∏î {round_str})\n"
                                    f"   - ‡∏ó‡∏∞‡πÄ‡∏ö‡∏µ‡∏¢‡∏ô: {job['Plate']}\n"
                                    f"   - ‡∏Ñ‡∏ô‡∏Ç‡∏±‡∏ö: {job['Driver']} ({phone})\n"
                                    f"   - ‚è≥ ‡∏™‡∏≤‡∏¢: {int(hours_late)} ‡∏ä‡∏°. {minutes_late} ‡∏ô‡∏≤‡∏ó‡∏µ")
                        
                        if is_day: late_list['day'].append(info_txt)
                        else: late_list['night'].append(info_txt)
            except Exception as e: 
                continue

        current_hour_key = now_thai.strftime("%Y-%m-%d_%H")

        if late_list['day']:
            notify_key = f"late_alert_day_{current_hour_key}"
            if not is_already_notified(sheet, notify_key):
                msg = f"‚ö†Ô∏è **‡πÄ‡∏ï‡∏∑‡∏≠‡∏ô‡∏†‡∏±‡∏¢: ‡∏£‡∏ñ‡πÄ‡∏Ç‡πâ‡∏≤‡∏ä‡πâ‡∏≤‡πÄ‡∏Å‡∏¥‡∏ô 2 ‡∏ä‡∏°. (‡∏£‡∏≠‡∏ö‡πÄ‡∏ä‡πâ‡∏≤)**\n(‡πÅ‡∏à‡πâ‡∏á‡πÄ‡∏ï‡∏∑‡∏≠‡∏ô‡∏õ‡∏£‡∏∞‡∏à‡∏≥‡∏ä‡∏±‡πà‡∏ß‡πÇ‡∏°‡∏á‡∏ó‡∏µ‡πà {now_thai.hour}:00)\n" + "\n".join(late_list['day'])
                send_discord_msg(msg)
            
        if late_list['night']:
            notify_key = f"late_alert_night_{current_hour_key}"
            if not is_already_notified(sheet, notify_key):
                msg = f"‚ö†Ô∏è **‡πÄ‡∏ï‡∏∑‡∏≠‡∏ô‡∏†‡∏±‡∏¢: ‡∏£‡∏ñ‡πÄ‡∏Ç‡πâ‡∏≤‡∏ä‡πâ‡∏≤‡πÄ‡∏Å‡∏¥‡∏ô 2 ‡∏ä‡∏°. (‡∏£‡∏≠‡∏ö‡∏î‡∏∂‡∏Å)**\n(‡πÅ‡∏à‡πâ‡∏á‡πÄ‡∏ï‡∏∑‡∏≠‡∏ô‡∏õ‡∏£‡∏∞‡∏à‡∏≥‡∏ä‡∏±‡πà‡∏ß‡πÇ‡∏°‡∏á‡∏ó‡∏µ‡πà {now_thai.hour}:00)\n" + "\n".join(late_list['night'])
                send_discord_msg(msg)

    except Exception as e:
        print(f"Late Check Error: {e}")

# ======================================================
# [FIXED] get_db Function with Retry Logic & ID Support
# ======================================================
def get_db():
    scope = ["https://spreadsheets.google.com/feeds", 'https://www.googleapis.com/auth/spreadsheets', "https://www.googleapis.com/auth/drive.file", "https://www.googleapis.com/auth/drive"]
    creds_json = os.environ.get('GSPREAD_CREDENTIALS')
    
    if not creds_json:
        if os.path.exists("credentials.json"): 
            creds = ServiceAccountCredentials.from_json_keyfile_name("credentials.json", scope)
        else: 
            return None
    else:
        creds_dict = json.loads(creds_json)
        creds = ServiceAccountCredentials.from_json_keyfile_dict(creds_dict, scope)
    
    client = gspread.authorize(creds)
    
    # Retry Logic for Google API 500/429 Errors
    max_retries = 3
    for attempt in range(max_retries):
        try:
            if SPREADSHEET_ID and len(SPREADSHEET_ID) > 10:
                return client.open_by_key(SPREADSHEET_ID)
            else:
                return client.open("DriverLogApp")
        except Exception as e:
            if attempt == max_retries - 1:
                print(f"Failed to connect to Google Sheet after {max_retries} attempts: {e}")
                raise e
            print(f"Google Sheet API Error (Attempt {attempt+1}/{max_retries}). Retrying...")
            time.sleep(2)
            
    return None

# [Updated Login Route with Better Error Handling]
@app.route('/manager_login', methods=['GET', 'POST'])
def manager_login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        try:
            sheet = get_db()
            users = get_cached_records(sheet, 'Users')
            for user in users:
                if str(user['Username']) == username and str(user['Password']) == password:
                    session['user'] = username
                    return redirect(url_for('manager_dashboard'))
            return render_template('login.html', error="‡∏ä‡∏∑‡πà‡∏≠‡∏ú‡∏π‡πâ‡πÉ‡∏ä‡πâ‡∏´‡∏£‡∏∑‡∏≠‡∏£‡∏´‡∏±‡∏™‡∏ú‡πà‡∏≤‡∏ô‡πÑ‡∏°‡πà‡∏ñ‡∏π‡∏Å‡∏ï‡πâ‡∏≠‡∏á")
        except Exception as e: 
            err_msg = str(e)
            if "<Response [200]>" in err_msg or "500" in err_msg:
                err_msg = "‡∏£‡∏∞‡∏ö‡∏ö Google Sheets ‡∏Ç‡∏±‡∏î‡∏Ç‡πâ‡∏≠‡∏á‡∏ä‡∏±‡πà‡∏ß‡∏Ñ‡∏£‡∏≤‡∏ß ‡∏Å‡∏£‡∏∏‡∏ì‡∏≤‡∏•‡∏≠‡∏á‡πÉ‡∏´‡∏°‡πà ‡∏´‡∏£‡∏∑‡∏≠‡∏ï‡∏±‡πâ‡∏á‡∏Ñ‡πà‡∏≤ Spreadsheet ID ‡πÉ‡∏ô‡πÇ‡∏Ñ‡πâ‡∏î"
            return render_template('login.html', error=f"Error: {err_msg}")
    return render_template('login.html')

@app.route('/manager')
def manager_dashboard():
    if 'user' not in session: return redirect(url_for('manager_login'))
    
    sheet = get_db()
    raw_jobs = get_cached_records(sheet, 'Jobs')
    drivers = get_cached_records(sheet, 'Drivers')

    date_filter = request.args.get('date_filter')
    now_thai = datetime.now() + timedelta(hours=7)
    today_date = now_thai.strftime("%Y-%m-%d")
    if not date_filter: date_filter = today_date

    filtered_jobs = [j for j in raw_jobs if str(j['PO_Date']).strip() == str(date_filter).strip()]
    def sort_key(j):
        try: c = int(str(j['Car_No']).strip())
        except: c = 99999
        return (str(j['PO_Date']), c, str(j['Round']))
    filtered_jobs = sorted(filtered_jobs, key=sort_key)
    
    # [UPDATED] Logic ‡∏Ñ‡∏≥‡∏ô‡∏ß‡∏ì‡πÄ‡∏ß‡∏•‡∏≤‡πÄ‡∏Ç‡πâ‡∏≤‡∏™‡∏≤‡∏¢‡πÅ‡∏ö‡∏ö‡∏•‡∏∞‡πÄ‡∏≠‡∏µ‡∏¢‡∏î (‡∏ä‡∏±‡πà‡∏ß‡πÇ‡∏°‡∏á/‡∏ô‡∏≤‡∏ó‡∏µ)
    for job in filtered_jobs:
        job['is_start_late'] = False
        job['delay_msg'] = ""
        
        t_plan_str = str(job.get('Round', '')).strip()
        t_act_str = str(job.get('T2_StartLoad', '')).strip()
        
        if t_plan_str and t_act_str:
            try:
                fmt_plan = "%H:%M" if len(t_plan_str) <= 5 else "%H:%M:%S"
                fmt_act = "%H:%M" if len(t_act_str) <= 5 else "%H:%M:%S"
                
                # ‡πÉ‡∏ä‡πâ‡∏õ‡∏µ‡∏™‡∏°‡∏°‡∏ï‡∏¥‡πÄ‡∏û‡∏∑‡πà‡∏≠‡πÄ‡∏õ‡∏£‡∏µ‡∏¢‡∏ö‡πÄ‡∏ó‡∏µ‡∏¢‡∏ö‡πÄ‡∏ß‡∏•‡∏≤
                t_plan = datetime.strptime(t_plan_str, fmt_plan).replace(year=2000, month=1, day=1)
                t_act = datetime.strptime(t_act_str, fmt_act).replace(year=2000, month=1, day=1)
                
                # ‡∏à‡∏±‡∏î‡∏Å‡∏≤‡∏£‡∏Å‡∏£‡∏ì‡∏µ‡∏Ç‡πâ‡∏≤‡∏°‡∏ß‡∏±‡∏ô (Night Shift)
                if t_plan.hour >= 18 and t_act.hour < 6: 
                    t_act += timedelta(days=1)
                elif t_plan.hour < 6 and t_act.hour >= 18: 
                    t_act -= timedelta(days=1)
                    
                if t_act > t_plan: 
                    job['is_start_late'] = True
                    diff = t_act - t_plan
                    total_seconds = diff.total_seconds()
                    hours = int(total_seconds // 3600)
                    minutes = int((total_seconds % 3600) // 60)
                    job['delay_msg'] = f"‡∏•‡πà‡∏≤‡∏ä‡πâ‡∏≤ {hours} ‡∏ä‡∏°. {minutes} ‡∏ô."
            except: pass

    jobs_by_trip_key = {}
    total_done_jobs = 0
    total_branches = len(filtered_jobs)
    trip_last_end_time = {} 
    grouped_jobs_for_stats = []
    current_group = []
    prev_key = None
    late_arrivals_by_po = {}
    total_late_cars = 0

    for job in filtered_jobs:
        curr_key = (str(job['PO_Date']), str(job['Car_No']), str(job['Round']), str(job['Driver']))
        if curr_key != prev_key and prev_key is not None:
            grouped_jobs_for_stats.append(current_group)
            current_group = []
        current_group.append(job)
        prev_key = curr_key
        
        trip_key = (str(job['PO_Date']), str(job['Car_No']), str(job['Round']))
        if trip_key not in jobs_by_trip_key: jobs_by_trip_key[trip_key] = []
        jobs_by_trip_key[trip_key].append(job)
        if job['Status'] == 'Done': total_done_jobs += 1
            
        if not job.get('T1_Enter') and job['Status'] != 'Done':
            try:
                load_date_str = job.get('Load_Date', job['PO_Date'])
                round_str = str(job['Round']).strip()
                plan_dt = datetime.strptime(f"{load_date_str} {round_str}", "%Y-%m-%d %H:%M") if ' ' in f"{load_date_str} {round_str}" else None
                if not plan_dt: plan_dt = datetime.strptime(f"{job['PO_Date']} {round_str}", "%Y-%m-%d %H:%M")
                
                if str(job.get('Load_Date', '')).strip() == '' or str(job.get('Load_Date')) == str(job['PO_Date']):
                    if 0 <= plan_dt.hour < 6: plan_dt += timedelta(days=1)

                if now_thai > plan_dt:
                    po_key = str(job['PO_Date'])
                    if po_key not in late_arrivals_by_po: late_arrivals_by_po[po_key] = []
                    diff = now_thai - plan_dt
                    hours = int(diff.total_seconds() // 3600)
                    mins = int((diff.total_seconds() % 3600) // 60)
                    job['late_duration'] = f"{hours} ‡∏ä‡∏°. {mins} ‡∏ô."
                    if not any(x['Car_No'] == job['Car_No'] for x in late_arrivals_by_po[po_key]):
                        late_arrivals_by_po[po_key].append(job)
                        total_late_cars += 1
            except: pass
            
    if current_group: grouped_jobs_for_stats.append(current_group)

    driver_stats = {}
    for group in grouped_jobs_for_stats:
        first = group[0]
        d_name = first['Driver']
        if not d_name: continue
        if d_name not in driver_stats: driver_stats[d_name] = {'total_trips': 0, 'rounds': []}
        driver_stats[d_name]['total_trips'] += 1
        driver_stats[d_name]['rounds'].append({
            'round': first['Round'], 'car_no': first['Car_No'], 'plate': first['Plate'],
            'branches': [j['Branch_Name'] for j in group],
            'status': 'Done' if all(j['Status'] == 'Done' for j in group) else 'Pending'
        })
    
    for d in driver_stats:
        driver_stats[d]['rounds'].sort(key=lambda x: int(str(x['car_no']).strip()) if str(x['car_no']).strip().isdigit() else 9999)

    busy_day_drivers = set()
    busy_night_drivers = set()
    for job in filtered_jobs:
        if str(job.get('Status', '')).lower() == 'cancel': continue
        d_name = job.get('Driver')
        try:
            h = int(str(job.get('Round', '')).split(':')[0])
            if 6 <= h <= 18: busy_day_drivers.add(d_name)
            else: busy_night_drivers.add(d_name)
        except: pass

    idle_drivers_day = []
    idle_drivers_night = []
    idle_drivers_hybrid = []
    idle_drivers_new = []

    for d in drivers:
        d_name = d.get('Name')
        is_busy_day = d_name in busy_day_drivers
        is_busy_night = d_name in busy_night_drivers
        if not is_busy_day and not is_busy_night: idle_drivers_hybrid.append(d)
        elif not is_busy_day: idle_drivers_day.append(d)
        elif not is_busy_night: idle_drivers_night.append(d)

    shift_status = {'day': {'total': 0, 'entered': 0, 'finished': 0, 'is_enter_complete': False, 'is_job_complete': False},
                    'night': {'total': 0, 'entered': 0, 'finished': 0, 'is_enter_complete': False, 'is_job_complete': False}}
    
    for trip_key, job_list in jobs_by_trip_key.items():
        active = [j for j in job_list if str(j.get('Status', '')).lower() != 'cancel']
        if not active: continue
        first = active[0]
        is_day, _ = get_shift_info(first.get('Round'))
        target = shift_status['day'] if is_day else shift_status['night']
        target['total'] += 1
        if str(first.get('T1_Enter', '')).strip(): target['entered'] += 1
        if all(j['Status'] == 'Done' for j in active): target['finished'] += 1

    for k in shift_status:
        if shift_status[k]['total'] > 0:
            if shift_status[k]['total'] == shift_status[k]['entered']: shift_status[k]['is_enter_complete'] = True
            if shift_status[k]['total'] == shift_status[k]['finished']: shift_status[k]['is_job_complete'] = True

    completed_trips = 0
    for trip_key, job_list in jobs_by_trip_key.items():
        if all(job['Status'] == 'Done' for job in job_list):
            completed_trips += 1
            trip_last_end_time[trip_key] = max([j['T8_EndJob'] for j in job_list if j['T8_EndJob']], default="")
        else:
            trip_last_end_time[trip_key] = ""
            
    total_trips = len(jobs_by_trip_key)
    total_running_jobs = total_branches - total_done_jobs

    line_data_day = []
    line_data_night = []
    for group in grouped_jobs_for_stats:
        first = group[0]
        round_str = str(first['Round']).strip()
        is_day, _ = get_shift_info(round_str)
        status_txt = "‡∏¢‡∏±‡∏á‡πÑ‡∏°‡πà‡∏ñ‡∏∂‡∏á‡∏Ñ‡∏•‡∏±‡∏á"
        if all(j['Status'] == 'Done' for j in group): status_txt = f"‡∏à‡∏ö‡∏á‡∏≤‡∏ô‡∏ó‡∏∏‡∏Å‡∏™‡∏≤‡∏Ç‡∏≤ ({group[-1].get('T8_EndJob', '')})"
        elif first.get('T6_Exit'): status_txt = f"‡∏≠‡∏≠‡∏Å‡πÇ‡∏£‡∏á‡∏á‡∏≤‡∏ô ({first.get('T6_Exit')})"
        elif first.get('T1_Enter'): status_txt = f"‡πÄ‡∏Ç‡πâ‡∏≤‡πÇ‡∏£‡∏á‡∏á‡∏≤‡∏ô ({first.get('T1_Enter')})"
        
        trip_data = {'round': round_str, 'car_no': first['Car_No'], 'plate': first['Plate'], 'driver': first['Driver'],
                     'branches': [j['Branch_Name'] for j in group], 'load_date': first.get('Load_Date', first['PO_Date']), 'latest_status': status_txt}
        if is_day: line_data_day.append(trip_data)
        else: line_data_night.append(trip_data)
    
    line_data_day.sort(key=lambda x: x['round'])
    line_data_night.sort(key=lambda x: (int(x['round'].split(':')[0]) + 24 if int(x['round'].split(':')[0]) < 6 else int(x['round'].split(':')[0])))

    all_dates = sorted(list(set([str(j['PO_Date']).strip() for j in raw_jobs])), reverse=True)
    try:
        curr = datetime.strptime(date_filter, "%Y-%m-%d")
        prev = (curr - timedelta(days=1)).strftime("%Y-%m-%d")
        next_d = (curr + timedelta(days=1)).strftime("%Y-%m-%d")
    except: prev, next_d = date_filter, date_filter

    return render_template('manager.html', 
                           jobs=filtered_jobs, drivers=drivers, all_dates=all_dates, 
                           total_trips=total_trips, completed_trips=completed_trips,
                           total_branches=total_branches, total_done_jobs=total_done_jobs,
                           total_running_jobs=total_running_jobs, now_time=now_thai.strftime("%H:%M"),
                           today_date=today_date, current_filter_date=date_filter,
                           prev_date=prev, next_date=next_d, trip_last_end_time=trip_last_end_time,
                           line_data_day=line_data_day, line_data_night=line_data_night,
                           late_arrivals_by_po=late_arrivals_by_po, total_late_cars=total_late_cars,
                           driver_stats=driver_stats, idle_drivers_day=idle_drivers_day,
                           idle_drivers_night=idle_drivers_night, idle_drivers_hybrid=idle_drivers_hybrid,
                           idle_drivers_new=idle_drivers_new, shift_status=shift_status)
                           
# ==========================================
# [UPDATED] Create Job Function
# ==========================================
@app.route('/create_job', methods=['POST'])
def create_job():
    if 'user' not in session: return redirect(url_for('manager_login'))
    sheet = get_db()
    ws = sheet.worksheet('Jobs')
    
    po_date = request.form['po_date']
    load_date = request.form['load_date']
    round_time = request.form['round_time']
    car_no = request.form['car_no']
    driver_name = request.form['driver_name']
    weight = request.form.get('weight', '')
    branches = request.form.getlist('branches') 
    
    # --- [‡∏™‡πà‡∏ß‡∏ô‡∏™‡∏≥‡∏Ñ‡∏±‡∏ç] ‡∏£‡∏±‡∏ö‡∏Ñ‡πà‡∏≤ PO List ‡∏à‡∏≤‡∏Å Textarea ---
    po_input_raw = request.form.get('po_list_input', '')
    po_str_to_save = ""
    if po_input_raw:
        # ‡πÅ‡∏¢‡∏Å‡∏ö‡∏£‡∏£‡∏ó‡∏±‡∏î ‡∏ï‡∏±‡∏î‡∏ä‡πà‡∏≠‡∏á‡∏ß‡πà‡∏≤‡∏á ‡πÅ‡∏•‡∏∞‡∏£‡∏ß‡∏°‡πÄ‡∏õ‡πá‡∏ô string ‡πÄ‡∏î‡∏µ‡∏¢‡∏ß‡∏Ñ‡∏±‡πà‡∏ô‡∏î‡πâ‡∏ß‡∏¢ comma
        po_lines = [line.strip() for line in po_input_raw.splitlines() if line.strip()]
        po_str_to_save = ",".join(po_lines)
    # ---------------------------------------------
    
    drivers_ws = sheet.worksheet('Drivers')
    driver_list = drivers_ws.get_all_records()
    plate = ""
    for d in driver_list:
        if d['Name'] == driver_name:
            plate = d['Plate_License']
            break
            
    new_rows = []
    for branch in branches:
        if branch.strip(): 
            # ‡∏™‡∏£‡πâ‡∏≤‡∏á‡πÅ‡∏ñ‡∏ß‡∏Ç‡πâ‡∏≠‡∏°‡∏π‡∏• ‡πÇ‡∏î‡∏¢‡πÉ‡∏™‡πà PO ‡∏•‡∏á‡πÉ‡∏ô Column Z (‡∏•‡∏≥‡∏î‡∏±‡∏ö‡∏ó‡∏µ‡πà 26)
            row = [
                po_date, load_date, round_time, car_no, driver_name, plate, branch, weight, 
                "", "", "", "", "", "", "", "", "New",  # ‡∏ñ‡∏∂‡∏á Column Q (17)
                "", "", "", "", "", "", "", "",         # Column R-Y (18-25)
                po_str_to_save,                         # Column Z (26) : PO_Nos
                "",                                     # Column AA (27) : Doc_Result
                ""                                      # Column AB (28) : Weight_Result
            ]
            new_rows.append(row)
    
    if new_rows: 
        ws.append_rows(new_rows)
        invalidate_cache('Jobs')
    
    return redirect(url_for('manager_dashboard'))

@app.route('/delete_job', methods=['POST'])
def delete_job():
    if 'user' not in session: return redirect(url_for('manager_login'))
    sheet = get_db()
    ws = sheet.worksheet('Jobs')
    
    po_date = request.form['po_date']
    round_time = request.form['round_time']
    car_no = request.form['car_no']
    
    try:
        all_values = ws.get_all_values()
        rows_to_delete = []
        
        for i, row in enumerate(all_values):
            if i > 0: 
                if (row[0] == po_date and 
                    str(row[2]) == str(round_time) and  
                    str(row[3]) == str(car_no)):        
                    rows_to_delete.append(i + 1)
                    
        for row_idx in sorted(rows_to_delete, reverse=True):
            ws.delete_rows(row_idx)
        
        invalidate_cache('Jobs')
            
        return redirect(url_for('manager_dashboard'))
    except Exception as e: return f"Error: {e}"

@app.route('/export_excel')
def export_excel():
    sheet = get_db()
    raw_jobs = get_cached_records(sheet, 'Jobs')
    
    date_filter = request.args.get('date_filter')
    if date_filter:
        jobs = [j for j in raw_jobs if str(j['PO_Date']).strip() == str(date_filter).strip()]
    else:
        jobs = raw_jobs
        
    def sort_key_func(job):
        po_date = str(job['PO_Date'])
        car_no_str = str(job['Car_No']).strip()
        round_val = str(job['Round'])
        try: car_no_int = int(car_no_str)
        except ValueError: car_no_int = 99999 
        return (po_date, car_no_int, round_val)

    jobs = sorted(jobs, key=sort_key_func)
    
    export_data = []
    prev_trip_key = None
    
    grouped_jobs_for_summary = []
    current_group = []
    
    for job_index, job in enumerate(jobs):
        current_trip_key = (str(job['PO_Date']), str(job['Car_No']), str(job['Round']), str(job['Driver']))
        is_same = (current_trip_key == prev_trip_key)
        
        if current_trip_key != prev_trip_key and prev_trip_key is not None:
            grouped_jobs_for_summary.append(current_group)
            current_group = []
        current_group.append(job)
        
        t2_display = job['T2_StartLoad']
        if not is_same and job['T2_StartLoad']: 
            try:
                plan_time_str = str(job['Round']).strip()
                actual_time_str = str(job['T2_StartLoad']).strip()
                if plan_time_str and actual_time_str:
                    fmt = "%H:%M" if len(plan_time_str) <= 5 else "%H:%M:%S"
                    fmt_act = "%H:%M" if len(actual_time_str) <= 5 else "%H:%M:%S"
                    t_plan = datetime.strptime(plan_time_str, fmt)
                    t_act = datetime.strptime(actual_time_str, fmt_act)
                    
                    if (t_plan - t_act).total_seconds() > 12 * 3600:
                        t_act = t_act + timedelta(days=1)
                    
                    if t_act > t_plan:
                        diff = t_act - t_plan
                        total_seconds = diff.total_seconds()
                        hours = int(total_seconds // 3600)
                        minutes = int((total_seconds % 3600) // 60)
                        delay_msg = f" (‡∏•‡πà‡∏≤‡∏ä‡πâ‡∏≤ {hours} ‡∏ä‡∏°. {minutes} ‡∏ô.)"
                        t2_display = f"{actual_time_str}{delay_msg}"
            except: pass 

        formatted_date = job['PO_Date']
        try:
            date_obj = datetime.strptime(str(job['PO_Date']).strip(), "%Y-%m-%d")
            formatted_date = date_obj.strftime("%d/%m/%Y")
        except: pass
            
        row = {
            '‡∏•‡∏≥‡∏î‡∏±‡∏ö‡∏£‡∏ñ': "" if is_same else job['Car_No'],
            'PO Date': "" if is_same else formatted_date,
            '‡πÄ‡∏ß‡∏•‡∏≤‡πÇ‡∏´‡∏•‡∏î': "" if is_same else job['Round'], 
            '‡∏Ñ‡∏ô‡∏Ç‡∏±‡∏ö': "" if is_same else job['Driver'],
            '‡∏õ‡∏•‡∏≤‡∏¢‡∏ó‡∏≤‡∏á (‡∏™‡∏≤‡∏Ç‡∏≤)': job['Branch_Name'],
            '‡∏ô‡πâ‡∏≥‡∏´‡∏ô‡∏±‡∏Å': "" if is_same else comma_format(job.get('Weight', '')),
            '‡∏ó‡∏∞‡πÄ‡∏ö‡∏µ‡∏¢‡∏ô‡∏£‡∏ñ': "" if is_same else job['Plate'],
            '‡πÄ‡∏Ç‡πâ‡∏≤‡πÇ‡∏£‡∏á‡∏á‡∏≤‡∏ô': "" if is_same else job['T1_Enter'],
            '‡πÄ‡∏£‡∏¥‡πà‡∏°‡πÇ‡∏´‡∏•‡∏î': "" if is_same else t2_display, 
            '‡πÇ‡∏´‡∏•‡∏î‡πÄ‡∏™‡∏£‡πá‡∏à': "" if is_same else job['T3_EndLoad'],
            '‡∏¢‡∏∑‡πà‡∏ô‡πÄ‡∏≠‡∏Å‡∏™‡∏≤‡∏£': "" if is_same else job['T4_SubmitDoc'],
            '‡∏£‡∏±‡∏ö‡πÄ‡∏≠‡∏Å‡∏™‡∏≤‡∏£': "" if is_same else job['T5_RecvDoc'],
            '‡∏≠‡∏≠‡∏Å‡πÇ‡∏£‡∏á‡∏á‡∏≤‡∏ô': "" if is_same else job['T6_Exit'],
            '‡∏ñ‡∏∂‡∏á‡∏™‡∏≤‡∏Ç‡∏≤': job['T7_ArriveBranch'],
            '‡∏à‡∏ö‡∏á‡∏≤‡∏ô': job['T8_EndJob']
        }
        export_data.append(row)
        prev_trip_key = current_trip_key

    if current_group: grouped_jobs_for_summary.append(current_group)

    df = pd.DataFrame(export_data)
    
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Report')
    
    output.seek(0)
    wb = load_workbook(output)
    ws = wb.active
    
    font_header = Font(name='Cordia New', size=14, bold=True, color='FFFFFF') 
    font_body = Font(name='Cordia New', size=14)
    font_summary_head = Font(name='Cordia New', size=14, bold=True)
    font_summary_body = Font(name='Cordia New', size=14)
    
    side_thin = Side(border_style="thin", color="000000")
    side_none = Side(border_style=None) 
    border_all = Border(top=side_thin, bottom=side_thin, left=side_thin, right=side_thin)
    border_header = Border(top=side_thin, bottom=side_thin, left=side_thin, right=side_thin)
    
    align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    align_left = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    fill_header = PatternFill(start_color='2E4053', end_color='2E4053', fill_type='solid')
    fill_white = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
    fill_blue_light = PatternFill(start_color='EBF5FB', end_color='EBF5FB', fill_type='solid')
    fill_green_branch = PatternFill(start_color='D5F5E3', end_color='D5F5E3', fill_type='solid')
    fill_red_end = PatternFill(start_color='FADBD8', end_color='FADBD8', fill_type='solid')
    fill_sum_head = PatternFill(start_color='D6EAF8', end_color='D6EAF8', fill_type='solid') 
    fill_sum_total = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid') 

    ws.freeze_panes = 'A2'

    current_trip_id = None
    is_zebra_active = False

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        ws.row_dimensions[row[0].row].height = 21

        if row[0].row == 1:
            for cell in row:
                cell.font = font_header
                cell.fill = fill_header
                cell.alignment = align_center
                cell.border = border_header
            continue

        job_index = row[0].row - 2
        is_group_end = False
        if 0 <= job_index < len(jobs):
            job_data = jobs[job_index]
            if job_index == len(jobs) - 1: is_group_end = True
            else:
                next_job = jobs[job_index + 1]
                if (str(job_data['PO_Date']) != str(next_job['PO_Date'])) or (str(job_data['Car_No']) != str(next_job['Car_No'])) or (str(job_data['Round']) != str(next_job['Round'])):
                    is_group_end = True

            this_trip_key = (str(job_data['PO_Date']), str(job_data['Car_No']), str(job_data['Round']))
            if this_trip_key != current_trip_id:
                is_zebra_active = not is_zebra_active
                current_trip_id = this_trip_key

        row_fill = fill_blue_light if is_zebra_active else fill_white
        current_border = Border(left=side_thin, right=side_thin, top=side_none, bottom=side_thin if is_group_end else side_none)

        for cell in row:
            col_name = ws.cell(row=1, column=cell.column).value
            f_bold = False
            f_color = '000000'
            if col_name in ['‡∏ñ‡∏∂‡∏á‡∏™‡∏≤‡∏Ç‡∏≤', '‡∏à‡∏ö‡∏á‡∏≤‡∏ô', '‡πÄ‡∏£‡∏¥‡πà‡∏°‡πÇ‡∏´‡∏•‡∏î']: f_bold = True
            
            if col_name == '‡πÄ‡∏£‡∏¥‡πà‡∏°‡πÇ‡∏´‡∏•‡∏î':
                cell_val_str = str(cell.value) if cell.value else ""
                if "(‡∏•‡πà‡∏≤‡∏ä‡πâ‡∏≤" in cell_val_str: f_color = 'C0392B'
                elif cell_val_str.strip() != "": f_color = '196F3D'

            cell.font = Font(name='Cordia New', size=14, bold=f_bold, color=f_color)
            cell.border = current_border 
            cell.fill = row_fill
            
            if col_name == '‡∏ñ‡∏∂‡∏á‡∏™‡∏≤‡∏Ç‡∏≤': cell.fill = fill_green_branch
            elif col_name == '‡∏à‡∏ö‡∏á‡∏≤‡∏ô': cell.fill = fill_red_end
            
            if col_name in ['‡∏Ñ‡∏ô‡∏Ç‡∏±‡∏ö', '‡∏õ‡∏•‡∏≤‡∏¢‡∏ó‡∏≤‡∏á (‡∏™‡∏≤‡∏Ç‡∏≤)', '‡∏ó‡∏∞‡πÄ‡∏ö‡∏µ‡∏¢‡∏ô‡∏£‡∏ñ']: cell.alignment = align_left
            else: cell.alignment = align_center

    def create_counter(): return {'count':0, 't1':0, 't2':0, 't3':0, 't4':0, 't5':0, 't6':0, 't7':0, 't8':0}
    sum_day = create_counter()
    sum_night = create_counter()
    
    for group in grouped_jobs_for_summary:
        if not group: continue
        first_job = group[0]
        last_job = group[-1]
        
        round_time = str(first_job.get('Round', '')).strip()
        is_day_shift = True
        try:
            hour = int(round_time.split(':')[0])
            if not (6 <= hour <= 18): is_day_shift = False
        except: pass

        target = sum_day if is_day_shift else sum_night
        target['count'] += 1 
        if first_job.get('T1_Enter'): target['t1'] += 1
        if first_job.get('T2_StartLoad'): target['t2'] += 1
        if first_job.get('T3_EndLoad'): target['t3'] += 1
        if first_job.get('T4_SubmitDoc'): target['t4'] += 1
        if first_job.get('T5_RecvDoc'): target['t5'] += 1
        if first_job.get('T6_Exit'): target['t6'] += 1
        if first_job.get('T7_ArriveBranch'): target['t7'] += 1 
        if last_job.get('T8_EndJob'): target['t8'] += 1       

    sum_total = create_counter()
    for k in sum_total: sum_total[k] = sum_day[k] + sum_night[k]

    start_row = ws.max_row + 2 
    # Adjusted columns for summary (shifted by 1 due to weight)
    summary_headers = ['‡∏£‡∏≠‡∏ö‡πÇ‡∏´‡∏•‡∏î', '‡∏à‡∏≥‡∏ô‡∏ß‡∏ô‡∏£‡∏ñ', '‡πÄ‡∏Ç‡πâ‡∏≤‡πÇ‡∏£‡∏á‡∏á‡∏≤‡∏ô', '‡πÄ‡∏£‡∏¥‡πà‡∏°‡πÇ‡∏´‡∏•‡∏î', '‡πÇ‡∏´‡∏•‡∏î‡πÄ‡∏™‡∏£‡πá‡∏à', '‡∏¢‡∏∑‡πà‡∏ô‡πÄ‡∏≠‡∏Å‡∏™‡∏≤‡∏£', '‡∏£‡∏±‡∏ö‡πÄ‡∏≠‡∏Å‡∏™‡∏≤‡∏£', '‡∏≠‡∏≠‡∏Å‡πÇ‡∏£‡∏á‡∏á‡∏≤‡∏ô', '‡∏ñ‡∏∂‡∏á‡∏™‡∏≤‡∏Ç‡∏≤', '‡∏à‡∏ö‡∏á‡∏≤‡∏ô']
    col_map_idx = [6, 7, 8, 9, 10, 11, 12, 13, 14, 15] 
    
    ws.cell(row=start_row, column=6, value="‡∏£‡∏≠‡∏ö‡πÇ‡∏´‡∏•‡∏î")
    for i, label in enumerate(summary_headers[1:]):
        ws.cell(row=start_row, column=col_map_idx[i+1], value=label)

    rows_to_write = [
        ('‡∏Å‡∏•‡∏≤‡∏á‡∏ß‡∏±‡∏ô', sum_day),
        ('‡∏Å‡∏•‡∏≤‡∏á‡∏Ñ‡∏∑‡∏ô', sum_night),
        ('‡∏£‡∏ß‡∏°', sum_total)
    ]
    
    for idx, (label, data) in enumerate(rows_to_write):
        curr_r = start_row + 1 + idx
        ws.row_dimensions[curr_r].height = 21
        ws.cell(row=curr_r, column=6, value=label)
        vals = [data['count'], data['t1'], data['t2'], data['t3'], data['t4'], data['t5'], data['t6'], data['t7'], data['t8']]
        for i, val in enumerate(vals):
            ws.cell(row=curr_r, column=col_map_idx[i+1], value=val)

    for r in range(start_row, start_row + 4):
        for c in range(6, 16): 
            cell = ws.cell(row=r, column=c)
            cell.border = border_all
            cell.alignment = align_center
            cell.font = font_summary_body
            if r == start_row:
                cell.fill = fill_sum_head
                cell.font = font_summary_head
            elif r == start_row + 3:
                cell.fill = fill_sum_total
                cell.font = font_summary_head

    for column_cells in ws.columns:
        col_letter = get_column_letter(column_cells[0].column)
        col_header = column_cells[0].value
        if col_header == '‡πÄ‡∏£‡∏¥‡πà‡∏°‡πÇ‡∏´‡∏•‡∏î': ws.column_dimensions[col_letter].width = 22.00 
        else:
            length = 0
            for cell in column_cells:
                if cell.row < start_row: 
                    val = str(cell.value) if cell.value else ""
                    lines = val.split('\n')
                    longest = max(len(line) for line in lines) if lines else 0
                    if longest > length: length = longest
            ws.column_dimensions[col_letter].width = min(length + 5, 50)

    final_output = io.BytesIO()
    wb.save(final_output)
    final_output.seek(0)
    filename = f"Report_{date_filter if date_filter else 'All'}.xlsx"
    return send_file(final_output, download_name=filename, as_attachment=True)
    
@app.route('/export_pdf')
def export_pdf():
    def thai_date_filter(date_str):
        try:
            d = datetime.strptime(date_str, "%Y-%m-%d")
            return f"{d.day} {['‡∏°.‡∏Ñ.','‡∏Å.‡∏û.','‡∏°‡∏µ.‡∏Ñ.','‡πÄ‡∏°.‡∏¢.','‡∏û.‡∏Ñ.','‡∏°‡∏¥.‡∏¢.','‡∏Å.‡∏Ñ.','‡∏™.‡∏Ñ.','‡∏Å.‡∏¢.','‡∏ï.‡∏Ñ.','‡∏û.‡∏¢.','‡∏ò.‡∏Ñ.'][d.month-1]} {d.year+543}"
        except: return date_str

    sheet = get_db()
    raw_jobs = get_cached_records(sheet, 'Jobs')
    date_filter = request.args.get('date_filter')
    if date_filter:
        jobs = [j for j in raw_jobs if str(j['PO_Date']).strip() == str(date_filter).strip()]
    else:
        jobs = raw_jobs
        
    def sort_key_func(job):
        po_date = str(job['PO_Date'])
        car_no_str = str(job['Car_No']).strip()
        round_val = str(job['Round'])
        try: car_no_int = int(car_no_str)
        except ValueError: car_no_int = 99999 
        return (po_date, car_no_int, round_val)

    jobs = sorted(jobs, key=sort_key_func)

    def create_counter(): return {'total': 0, 't1': 0, 't2': 0, 't3': 0, 't6': 0, 't7': 0, 't8': 0}
    sum_day = create_counter()
    sum_night = create_counter()
    grouped_jobs = []
    current_group = []
    prev_key = None
    
    for job in jobs:
        curr_key = (str(job['PO_Date']), str(job['Car_No']), str(job['Round']), str(job['Driver']))
        if curr_key != prev_key and prev_key is not None:
            grouped_jobs.append(current_group)
            current_group = []
        current_group.append(job)
        prev_key = curr_key
        
        job['is_late'] = False
        job['delay_msg'] = ""
        t_plan_str = str(job['Round']).strip()
        t_act_str = str(job['T2_StartLoad']).strip()
        
        if t_plan_str and t_act_str:
            try:
                fmt_plan = "%H:%M" if len(t_plan_str) <= 5 else "%H:%M:%S"
                fmt_act = "%H:%M" if len(t_act_str) <= 5 else "%H:%M:%S"
                t_plan = datetime.strptime(t_plan_str, fmt_plan)
                t_act = datetime.strptime(t_act_str, fmt_act)
                if (t_plan - t_act).total_seconds() > 12 * 3600: t_act = t_act + timedelta(days=1)
                if t_act > t_plan:
                    job['is_late'] = True
                    diff = t_act - t_plan
                    total_seconds = diff.total_seconds()
                    hours = int(total_seconds // 3600)
                    minutes = int((total_seconds % 3600) // 60)
                    job['delay_msg'] = f"(‡∏•‡πà‡∏≤‡∏ä‡πâ‡∏≤ {hours} ‡∏ä‡∏°. {minutes} ‡∏ô.)"
            except: pass
            
    if current_group: grouped_jobs.append(current_group)

    for group in grouped_jobs:
        if not group: continue
        first_job = group[0]
        round_time = str(first_job.get('Round', '')).strip()
        is_day_shift = True
        try:
            hour = int(round_time.split(':')[0])
            if not (6 <= hour <= 18): is_day_shift = False
        except: pass

        target_sum = sum_day if is_day_shift else sum_night
        target_sum['total'] += 1
        if first_job.get('T1_Enter'): target_sum['t1'] += 1
        if first_job.get('T2_StartLoad'): target_sum['t2'] += 1
        if first_job.get('T3_EndLoad'): target_sum['t3'] += 1
        if first_job.get('T6_Exit'): target_sum['t6'] += 1
        
        if any(str(j.get('T7_ArriveBranch', '')).strip() != '' for j in group): target_sum['t7'] += 1
        if all(str(j.get('T8_EndJob', '')).strip() != '' for j in group): target_sum['t8'] += 1

    sum_total = create_counter()
    for key in sum_total:
        sum_total[key] = sum_day[key] + sum_night[key]

    basedir = os.path.abspath(os.path.dirname(__file__))
    font_path = os.path.join(basedir, 'static', 'fonts', 'Sarabun-Regular.ttf')
    logo_path = os.path.join(basedir, 'static', 'mylogo.png') 
    po_date_thai = thai_date_filter(date_filter) if date_filter else "‡∏ó‡∏±‡πâ‡∏á‡∏´‡∏°‡∏î"
    print_date = (datetime.now() + timedelta(hours=7)).strftime("%d/%m/%Y %H:%M")

    class PDF(FPDF):
        def __init__(self, **kwargs):
            super().__init__(**kwargs)
            self.is_summary_page = False

        def header(self):
            self.add_font('Sarabun', '', font_path, uni=True)
            if self.is_summary_page:
                self.set_font('Sarabun', '', 18)
                self.set_y(25)
                self.cell(0, 15, f'‡∏™‡∏£‡∏∏‡∏õ‡∏†‡∏≤‡∏û‡∏£‡∏ß‡∏°‡∏Å‡∏≤‡∏£‡∏à‡∏±‡∏î‡∏™‡πà‡∏á‡∏™‡∏¥‡∏ô‡∏Ñ‡πâ‡∏≤ ‡∏õ‡∏£‡∏∞‡∏à‡∏≥‡∏ß‡∏±‡∏ô‡∏ó‡∏µ‡πà {po_date_thai}', align='C', new_x="LMARGIN", new_y="NEXT")
                self.ln(5)
            else:
                if os.path.exists(logo_path):
                    self.image(logo_path, x=7, y=8, w=18)
                
                self.set_font('Sarabun', '', 16) 
                self.set_y(10)
                self.cell(0, 8, '‡∏£‡∏≤‡∏¢‡∏á‡∏≤‡∏ô‡∏™‡∏£‡∏∏‡∏õ‡∏Å‡∏≤‡∏£‡∏à‡∏±‡∏î‡∏™‡πà‡∏á‡∏™‡∏¥‡∏ô‡∏Ñ‡πâ‡∏≤ (Daily Jobs Report)', align='C', new_x="LMARGIN", new_y="NEXT")
                self.set_font_size(14)
                self.cell(0, 8, '‡∏ö‡∏£‡∏¥‡∏©‡∏±‡∏ó ‡πÅ‡∏≠‡∏•‡πÄ‡∏≠‡πá‡∏°‡∏ó‡∏µ. ‡∏ó‡∏£‡∏≤‡∏ô‡∏™‡∏õ‡∏≠‡∏£‡πå‡∏ï ‡∏à‡∏≥‡∏Å‡∏±‡∏î', align='C', new_x="LMARGIN", new_y="NEXT")
                self.set_font_size(10)
                self.cell(0, 6, f'‡∏ß‡∏±‡∏ô‡∏ó‡∏µ‡πà‡πÄ‡∏≠‡∏Å‡∏™‡∏≤‡∏£: {po_date_thai} | ‡∏û‡∏¥‡∏°‡∏û‡πå‡πÄ‡∏°‡∏∑‡πà‡∏≠: {print_date}', align='C', new_x="LMARGIN", new_y="NEXT")
                self.ln(4)

                cols = [10, 30, 35, 16, 45, 12, 16, 35, 16, 16, 22, 22]
                headers = ['‡∏Ñ‡∏±‡∏ô‡∏ó‡∏µ‡πà', '‡∏ó‡∏∞‡πÄ‡∏ö‡∏µ‡∏¢‡∏ô', '‡∏Ñ‡∏ô‡∏Ç‡∏±‡∏ö', '‡πÄ‡∏ß‡∏•‡∏≤‡πÇ‡∏´‡∏•‡∏î', '‡∏õ‡∏•‡∏≤‡∏¢‡∏ó‡∏≤‡∏á', '‡∏ô‡∏ô.', '‡πÄ‡∏Ç‡πâ‡∏≤‡πÇ‡∏£‡∏á‡∏á‡∏≤‡∏ô', '‡πÄ‡∏£‡∏¥‡πà‡∏°‡πÇ‡∏´‡∏•‡∏î', '‡πÇ‡∏´‡∏•‡∏î‡πÄ‡∏™‡∏£‡πá‡∏à', '‡∏≠‡∏≠‡∏Å‡πÇ‡∏£‡∏á‡∏á‡∏≤‡∏ô', '‡∏ñ‡∏∂‡∏á‡∏™‡∏≤‡∏Ç‡∏≤', '‡∏à‡∏ö‡∏á‡∏≤‡∏ô']
                self.set_fill_color(44, 62, 80)
                self.set_text_color(255, 255, 255)
                self.set_font('Sarabun', '', 9) 
                for i, h in enumerate(headers):
                    self.cell(cols[i], 8, h, border=1, align='C', fill=True)
                self.ln()
                self.set_text_color(0, 0, 0)

        def footer(self):
            self.set_y(-15)
            self.set_font('Sarabun', '', 8)
            self.set_text_color(100, 100, 100)
            self.cell(0, 10, '‡∏Ç‡πâ‡∏≠‡∏°‡∏π‡∏•‡∏à‡∏≤‡∏Å: ‡∏£‡∏∞‡∏ö‡∏ö LMT. Transport Driver App V.1.02', align='L')
            self.set_x(-30)
            self.cell(0, 10, f'‡∏´‡∏ô‡πâ‡∏≤ {self.page_no()}/{{nb}}', align='R')

    pdf = PDF(orientation='L', unit='mm', format='A4')
    pdf.alias_nb_pages()
    pdf.set_margins(7, 10, 7)
    pdf.add_page()
    
    cols = [10, 30, 35, 16, 45, 12, 16, 35, 16, 16, 22, 22]
    
    for group in grouped_jobs:
        group_total_height = 0
        for job in group:
            h = 9
            if job.get('is_late', False): h = 13
            group_total_height += h

        if group_total_height > (pdf.page_break_trigger - pdf.get_y()):
            pdf.add_page()

        for idx, job in enumerate(group):
            is_first_row = (idx == 0)
            is_last_in_group = (idx == len(group) - 1)
            y_top = pdf.get_y()
            pdf.set_fill_color(255, 255, 255)
            
            c_no = str(job['Car_No']) if is_first_row else ""
            plate = str(job['Plate']) if is_first_row else ""
            driver = str(job['Driver']) if is_first_row else ""
            round_t = str(job['Round']) if is_first_row else ""
            branch = str(job['Branch_Name'])
            weight = comma_format(job.get('Weight', '')) if is_first_row else ""
            t1 = str(job['T1_Enter']) if is_first_row else ""
            
            t2_text = ""
            is_late_row = False
            if is_first_row:
                t2_text = str(job['T2_StartLoad'])
                if job['is_late']:
                    t2_text += f"\n{job['delay_msg']}"
                    is_late_row = True
            
            t3 = str(job['T3_EndLoad']) if is_first_row else ""
            t6 = str(job['T6_Exit']) if is_first_row else ""
            t7 = str(job['T7_ArriveBranch'])
            t8 = str(job['T8_EndJob'])

            row_height = 9
            if is_late_row: row_height = 13

            pdf.set_font('Sarabun', '', 8)
            pdf.set_text_color(0, 0, 0)
            pdf.cell(cols[0], row_height, c_no, border='LR', align='C')
            pdf.cell(cols[1], row_height, plate, border='LR', align='C')
            pdf.cell(cols[2], row_height, driver, border='LR', align='L')
            pdf.cell(cols[3], row_height, round_t, border='LR', align='C')
            
            pdf.set_font_size(7) 
            pdf.cell(cols[4], row_height, branch, border='LR', align='L')
            pdf.set_font_size(8) 
            pdf.cell(cols[5], row_height, weight, border='LR', align='C')
            
            pdf.cell(cols[6], row_height, t1, border='LR', align='C')

            current_x = pdf.get_x()
            current_y = pdf.get_y()
            if is_late_row:
                pdf.set_text_color(192, 57, 43)
                pdf.multi_cell(cols[7], row_height/2 if '\n' in t2_text else row_height, t2_text, border='LR', align='C')
                pdf.set_xy(current_x + cols[7], current_y)
                pdf.set_text_color(0, 0, 0)
            else:
                if is_first_row and t2_text: pdf.set_text_color(25, 111, 61)
                pdf.cell(cols[7], row_height, t2_text.split('\n')[0], border='LR', align='C')
                pdf.set_text_color(0, 0, 0)

            pdf.cell(cols[8], row_height, t3, border='LR', align='C')
            pdf.cell(cols[9], row_height, t6, border='LR', align='C')
            
            pdf.set_fill_color(213, 245, 227)
            pdf.cell(cols[10], row_height, t7, border='LR', align='C', fill=True)
            pdf.set_fill_color(250, 219, 216)
            pdf.cell(cols[11], row_height, t8, border='LR', align='C', fill=True)

            pdf.ln()
            
            if is_first_row:
                pdf.set_draw_color(0, 0, 0)
                pdf.set_line_width(0.3)
            else:
                pdf.set_draw_color(200, 200, 200)
                pdf.set_line_width(0.1)
                
            pdf.line(7, y_top, 282, y_top)

            if is_last_in_group:
                y_bottom = pdf.get_y()
                pdf.set_draw_color(0, 0, 0)
                pdf.set_line_width(0.3)
                pdf.line(7, y_bottom, 282, y_bottom)

            pdf.set_draw_color(0, 0, 0)
            pdf.set_line_width(0.2)

    pdf.is_summary_page = True
    pdf.add_page()
    
    sum_headers = ['‡∏£‡∏≠‡∏ö‡∏á‡∏≤‡∏ô', '‡∏à‡∏≥‡∏ô‡∏ß‡∏ô', '‡πÄ‡∏Ç‡πâ‡∏≤‡πÇ‡∏£‡∏á‡∏á‡∏≤‡∏ô', '‡πÄ‡∏Ç‡πâ‡∏≤‡πÇ‡∏´‡∏•‡∏î', '‡πÇ‡∏´‡∏•‡∏î‡πÄ‡∏™‡∏£‡πá‡∏à', '‡∏¢‡∏∑‡πà‡∏ô‡πÄ‡∏≠‡∏Å‡∏™‡∏≤‡∏£', '‡∏£‡∏±‡∏ö‡πÄ‡∏≠‡∏Å‡∏™‡∏≤‡∏£', '‡∏≠‡∏≠‡∏Å‡πÇ‡∏£‡∏á‡∏á‡∏≤‡∏ô', '‡∏ñ‡∏∂‡∏á‡∏™‡∏≤‡∏Ç‡∏≤', '‡∏à‡∏ö‡∏á‡∏≤‡∏ô']
    sum_cols = [[45, 25, 25, 25, 25, 25, 25, 25, 25, 25]] 
    total_table_width = sum(sum_cols)
    start_x = (297 - total_table_width) / 2 
    
    COLOR_HEADER_BG = (44, 62, 80)
    COLOR_HEADER_TXT = (255, 255, 255)
    COLOR_ROW_DAY_BG = (255, 255, 255)
    COLOR_ROW_NIGHT_BG = (242, 243, 244)
    COLOR_TOTAL_BG = (213, 245, 227)
    COLOR_TOTAL_TXT = (25, 111, 61)
    COLOR_BORDER = (189, 195, 199)

    def draw_sum_row(label, data, row_type='normal'):
        pdf.set_x(start_x)
        if row_type == 'header':
            pdf.set_fill_color(*COLOR_HEADER_BG)
            pdf.set_text_color(*COLOR_HEADER_TXT)
            pdf.set_font('Sarabun', '', 12)
            pdf.set_draw_color(*COLOR_HEADER_BG)
        elif row_type == 'total':
            pdf.set_fill_color(*COLOR_TOTAL_BG)
            pdf.set_text_color(*COLOR_TOTAL_TXT)
            pdf.set_font('Sarabun', '', 12)
            pdf.set_draw_color(*COLOR_BORDER)
        elif row_type == 'night':
            pdf.set_fill_color(*COLOR_ROW_NIGHT_BG)
            pdf.set_text_color(0, 0, 0)
            pdf.set_font('Sarabun', '', 11)
            pdf.set_draw_color(*COLOR_BORDER)
        else:
            pdf.set_fill_color(*COLOR_ROW_DAY_BG)
            pdf.set_text_color(0, 0, 0)
            pdf.set_font('Sarabun', '', 11)
            pdf.set_draw_color(*COLOR_BORDER)

        row_h = 12
        pdf.cell(sum_cols[0], row_h, label, border=1, align='C', fill=True)
        
        vals = []
        if row_type == 'header': vals = data
        else: vals = [str(data['total']), str(data['t1']), str(data['t2']), str(data['t3']), str(data['t6']), str(data['t7']), str(data['t8'])]

        for i, val in enumerate(vals):
            pdf.cell(sum_cols[i+1], row_h, val, border=1, align='C', fill=True)
        pdf.ln()

    draw_sum_row("‡∏£‡∏≠‡∏ö‡∏á‡∏≤‡∏ô", sum_headers[1:], row_type='header')
    draw_sum_row("‡∏£‡∏≠‡∏ö‡∏Å‡∏•‡∏≤‡∏á‡∏ß‡∏±‡∏ô", sum_day, row_type='day')
    draw_sum_row("‡∏£‡∏≠‡∏ö‡∏Å‡∏•‡∏≤‡∏á‡∏Ñ‡∏∑‡∏ô", sum_night, row_type='night')
    draw_sum_row("‡∏£‡∏ß‡∏°‡∏ó‡∏±‡πâ‡∏á‡∏´‡∏°‡∏î", sum_total, row_type='total')
    
    pdf.ln(5)
    pdf.set_x(start_x)
    pdf.set_text_color(127, 140, 141)
    pdf.set_font('Sarabun', '', 9)
    pdf.cell(0, 5, "* ‡∏Ç‡πâ‡∏≠‡∏°‡∏π‡∏•‡∏ô‡∏±‡∏ö‡∏à‡∏≤‡∏Å‡∏à‡∏≥‡∏ô‡∏ß‡∏ô‡πÄ‡∏ó‡∏µ‡πà‡∏¢‡∏ß‡∏£‡∏ñ‡∏ó‡∏µ‡πà‡∏°‡∏µ‡∏Å‡∏≤‡∏£‡∏ö‡∏±‡∏ô‡∏ó‡∏∂‡∏Å‡πÄ‡∏ß‡∏•‡∏≤‡πÉ‡∏ô‡πÅ‡∏ï‡πà‡∏•‡∏∞‡∏Ç‡∏±‡πâ‡∏ô‡∏ï‡∏≠‡∏ô‡∏à‡∏£‡∏¥‡∏á", align='L')

    pdf_bytes = pdf.output()
    filename = f"Summary_{date_filter if date_filter else 'All'}.pdf"
    return send_file(io.BytesIO(pdf_bytes), mimetype='application/pdf', as_attachment=True, download_name=filename) 

@app.route('/export_pdf_summary')
def export_pdf_summary():
    # --- Helper function for PDF: thai_date_filter is missing from the provided code, mocking it to prevent crash ---
    def thai_date_filter(date_str):
        try:
            d = datetime.strptime(date_str, "%Y-%m-%d")
            return f"{d.day} {['‡∏°.‡∏Ñ.','‡∏Å.‡∏û.','‡∏°‡∏µ.‡∏Ñ.','‡πÄ‡∏°.‡∏¢.','‡∏û.‡∏Ñ.','‡∏°‡∏¥.‡∏¢.','‡∏Å.‡∏Ñ.','‡∏™.‡∏Ñ.','‡∏Å.‡∏¢.','‡∏ï.‡∏Ñ.','‡∏û.‡∏¢.','‡∏ò.‡∏Ñ.'][d.month-1]} {d.year+543}"
        except: return date_str
    # -----------------------------------------------------------------------------------------------------------------

    sheet = get_db()
    raw_jobs = get_cached_records(sheet, 'Jobs')
    
    date_filter = request.args.get('date_filter')
    if date_filter:
        jobs = [j for j in raw_jobs if str(j['PO_Date']).strip() == str(date_filter).strip()]
    else:
        jobs = raw_jobs
        
    def sort_key_func(job):
        po_date = str(job['PO_Date'])
        car_no_str = str(job['Car_No']).strip()
        round_val = str(job['Round'])
        try: car_no_int = int(car_no_str)
        except ValueError: car_no_int = 99999 
        return (po_date, car_no_int, round_val)

    jobs = sorted(jobs, key=sort_key_func)

    def create_counter(): return {'count':0, 't1':0, 't2':0, 't3':0, 't4':0, 't5':0, 't6':0, 't7':0, 't8':0}
    sum_day = create_counter()
    sum_night = create_counter()
    grouped_jobs = []
    current_group = []
    prev_key = None

    for job in jobs:
        curr_key = (str(job['PO_Date']), str(job['Car_No']), str(job['Round']), str(job['Driver']))
        if curr_key != prev_key and prev_key is not None:
            grouped_jobs.append(current_group)
            current_group = []
        current_group.append(job)
        prev_key = curr_key
        
        job['is_late'] = False
        t_plan_str = str(job['Round']).strip()
        t_act_str = str(job['T2_StartLoad']).strip()
        if t_plan_str and t_act_str:
            try:
                fmt_plan = "%H:%M" if len(t_plan_str) <= 5 else "%H:%M:%S"
                fmt_act = "%H:%M" if len(t_act_str) <= 5 else "%H:%M:%S"
                t_plan = datetime.strptime(t_plan_str, fmt_plan)
                t_act = datetime.strptime(t_act_str, fmt_act)
                if (t_plan - t_act).total_seconds() > 12 * 3600: t_act += timedelta(days=1)
                if t_act > t_plan: job['is_late'] = True
            except: pass
            
    if current_group: grouped_jobs.append(current_group)

    for group in grouped_jobs:
        if not group: continue
        first_job = group[0]
        
        round_time = str(first_job.get('Round', '')).strip()
        is_day_shift = True
        try:
            hour = int(round_time.split(':')[0])
            if not (6 <= hour <= 18): is_day_shift = False
        except: pass

        target = sum_day if is_day_shift else sum_night
        target['count'] += 1 
        if first_job.get('T1_Enter'): target['t1'] += 1
        if first_job.get('T2_StartLoad'): target['t2'] += 1
        if first_job.get('T3_EndLoad'): target['t3'] += 1
        if first_job.get('T4_SubmitDoc'): target['t4'] += 1
        if first_job.get('T5_RecvDoc'): target['t5'] += 1
        if first_job.get('T6_Exit'): target['t6'] += 1
        
        if any(str(j.get('T7_ArriveBranch', '')).strip() != '' for j in group): target['t7'] += 1 
        if all(str(j.get('T8_EndJob', '')).strip() != '' for j in group): target['t8'] += 1       

    sum_total = create_counter()
    for k in sum_total: sum_total[k] = sum_day[k] + sum_night[k]

    basedir = os.path.abspath(os.path.dirname(__file__))
    font_path = os.path.join(basedir, 'static', 'fonts', 'Sarabun-Regular.ttf')
    logo_path = os.path.join(basedir, 'static', 'mylogo.png') 
    po_date_thai = thai_date_filter(date_filter) if date_filter else "‡∏ó‡∏±‡πâ‡∏á‡∏´‡∏°‡∏î"
    print_date = (datetime.now() + timedelta(hours=7)).strftime("%d/%m/%Y %H:%M")

    COLS = [10, 18, 27, 13, 50, 13, 13, 13, 13, 13, 13]
    HEADERS = ['‡∏Ñ‡∏±‡∏ô‡∏ó‡∏µ‡πà', '‡∏ó‡∏∞‡πÄ‡∏ö‡∏µ‡∏¢‡∏ô', '‡∏Ñ‡∏ô‡∏Ç‡∏±‡∏ö', '‡πÄ‡∏ß‡∏•‡∏≤‡πÇ‡∏´‡∏•‡∏î', '‡∏õ‡∏•‡∏≤‡∏¢‡∏ó‡∏≤‡∏á', '‡πÄ‡∏Ç‡πâ‡∏≤‡πÇ‡∏£‡∏á‡∏á‡∏≤‡∏ô', '‡πÄ‡∏£‡∏¥‡πà‡∏°‡πÇ‡∏´‡∏•‡∏î', '‡πÇ‡∏´‡∏•‡∏î‡πÄ‡∏™‡∏£‡πá‡∏à', '‡∏≠‡∏≠‡∏Å‡πÇ‡∏£‡∏á‡∏á‡∏≤‡∏ô', '‡∏ñ‡∏∂‡∏á‡∏™‡∏≤‡∏Ç‡∏≤', '‡∏à‡∏ö‡∏á‡∏≤‡∏ô']

    class PDFSummary(FPDF):
        def header(self):
            self.add_font('Sarabun', '', font_path, uni=True)
            if os.path.exists(logo_path):
                self.image(logo_path, x=7, y=6, w=10)
            self.set_font('Sarabun', '', 12) 
            self.set_y(6)
            self.cell(0, 8, '‡∏™‡∏£‡∏∏‡∏õ‡∏£‡∏≤‡∏¢‡∏á‡∏≤‡∏ô‡∏Å‡∏≤‡∏£‡∏à‡∏±‡∏î‡∏™‡πà‡∏á‡∏™‡∏¥‡∏ô‡∏Ñ‡πâ‡∏≤ (Compact View)', align='C', new_x="LMARGIN", new_y="NEXT")
            self.set_font_size(10)
            self.cell(0, 7, '‡∏ö‡∏£‡∏¥‡∏©‡∏±‡∏ó ‡πÅ‡∏≠‡∏•‡πÄ‡∏≠‡πá‡∏°‡∏ó‡∏µ. ‡∏ó‡∏£‡∏≤‡∏ô‡∏™‡∏õ‡∏≠‡∏£‡πå‡∏ï ‡∏à‡∏≥‡∏Å‡∏±‡∏î', align='C', new_x="LMARGIN", new_y="NEXT")
            self.set_font_size(8)
            self.cell(0, 6, f'‡∏ß‡∏±‡∏ô‡∏ó‡∏µ‡πà‡πÄ‡∏≠‡∏Å‡∏™‡∏≤‡∏£: {po_date_thai} | ‡∏û‡∏¥‡∏°‡∏û‡πå‡πÄ‡∏°‡∏∑‡πà‡∏≠: {print_date}', align='C', new_x="LMARGIN", new_y="NEXT")
            self.ln(3)
            self.set_fill_color(44, 62, 80)
            self.set_text_color(255, 255, 255)
            self.set_draw_color(100, 100, 100)
            self.set_font('Sarabun', '', 7)
            for i, h in enumerate(HEADERS):
                self.cell(COLS[i], 7, h, border=1, align='C', fill=True)
            self.ln()
            self.set_text_color(0, 0, 0)
            self.set_draw_color(200, 200, 200)

        def footer(self):
            self.set_y(-10)
            self.set_font('Sarabun', '', 6)
            self.set_text_color(150)
            self.cell(0, 10, f'‡∏´‡∏ô‡πâ‡∏≤ {self.page_no()}/{{nb}}', align='R')

    pdf = PDFSummary(orientation='P', unit='mm', format='A4')
    pdf.alias_nb_pages()
    pdf.set_margins(7, 7, 7)
    pdf.add_page()
    
    group_count = 0
    for group in grouped_jobs:
        if group_count % 2 == 0: pdf.set_fill_color(255, 255, 255) 
        else: pdf.set_fill_color(245, 247, 249) 
        group_count += 1
        
        for idx, job in enumerate(group):
            is_first_row = (idx == 0)
            is_last_in_group = (idx == len(group) - 1)
            
            c_no = str(job['Car_No']) if is_first_row else ""
            plate = str(job['Plate']) if is_first_row else ""
            driver = str(job['Driver']) if is_first_row else ""
            round_t = str(job['Round']) if is_first_row else ""
            branch = str(job['Branch_Name'])
            t1 = str(job['T1_Enter']) if is_first_row else ""
            
            t2 = ""
            is_late_row = False
            if is_first_row:
                t2 = str(job['T2_StartLoad'])
                if job['is_late']: is_late_row = True

            t3 = str(job['T3_EndLoad']) if is_first_row else ""
            t6 = str(job['T6_Exit']) if is_first_row else ""
            t7 = str(job['T7_ArriveBranch'])
            t8 = str(job['T8_EndJob'])

            row_height = 5.8
            if pdf.get_y() + row_height > pdf.page_break_trigger:
                pdf.add_page()
                if (group_count - 1) % 2 == 0: pdf.set_fill_color(255, 255, 255)
                else: pdf.set_fill_color(245, 247, 249)

            pdf.set_font('Sarabun', '', 7)
            pdf.set_draw_color(180, 180, 180) 
            
            pdf.cell(COLS[0], row_height, c_no, border=1, align='C', fill=True)
            pdf.cell(COLS[1], row_height, plate, border=1, align='C', fill=True)
            if pdf.get_string_width(driver) > COLS[2] - 2:
                 while pdf.get_string_width(driver + "..") > COLS[2] - 2 and len(driver) > 0:
                     driver = driver[:-1]
                 driver += ".."
            pdf.cell(COLS[2], row_height, driver, border=1, align='L', fill=True)
            pdf.cell(COLS[3], row_height, round_t, border=1, align='C', fill=True)
            
            if pdf.get_string_width(branch) > COLS[4] - 2:
                 while pdf.get_string_width(branch + "..") > COLS[4] - 2 and len(branch) > 0:
                     branch = branch[:-1]
                 branch += ".."
            pdf.cell(COLS[4], row_height, branch, border=1, align='L', fill=True)
            
            pdf.cell(COLS[5], row_height, t1, border=1, align='C', fill=True)

            if is_late_row: pdf.set_text_color(192, 57, 43) 
            elif t2: pdf.set_text_color(39, 174, 96)
            pdf.cell(COLS[6], row_height, t2, border=1, align='C', fill=True)
            pdf.set_text_color(0, 0, 0)

            pdf.cell(COLS[7], row_height, t3, border=1, align='C', fill=True)
            pdf.cell(COLS[8], row_height, t6, border=1, align='C', fill=True)
            
            base_r, base_g, base_b = (255, 255, 255) if (group_count - 1) % 2 == 0 else (245, 247, 249)
            pdf.set_fill_color(240, 253, 244)
            pdf.cell(COLS[9], row_height, t7, border=1, align='C', fill=True)
            pdf.set_fill_color(254, 242, 242)
            pdf.cell(COLS[10], row_height, t8, border=1, align='C', fill=True)
            
            pdf.set_fill_color(base_r, base_g, base_b)
            pdf.ln()
            
            if is_last_in_group:
                y_curr = pdf.get_y()
                pdf.set_draw_color(0, 0, 0)
                pdf.set_line_width(0.3)
                pdf.line(7, y_curr, 203, y_curr) 
                pdf.set_line_width(0.2)
                pdf.set_draw_color(180, 180, 180)

    pdf.ln(5)
    if pdf.get_y() + 30 > pdf.page_break_trigger:
        pdf.add_page()
    
    SUM_HEADERS = ['‡∏£‡∏≠‡∏ö‡∏á‡∏≤‡∏ô', '‡∏à‡∏≥‡∏ô‡∏ß‡∏ô', '‡πÄ‡∏Ç‡πâ‡∏≤‡πÇ‡∏£‡∏á‡∏á‡∏≤‡∏ô', '‡πÄ‡∏£‡∏¥‡πà‡∏°‡πÇ‡∏´‡∏•‡∏î', '‡πÇ‡∏´‡∏•‡∏î‡πÄ‡∏™‡∏£‡πá‡∏à', '‡∏¢‡∏∑‡πà‡∏ô‡πÄ‡∏≠‡∏Å‡∏™‡∏≤‡∏£', '‡∏£‡∏±‡∏ö‡πÄ‡∏≠‡∏Å‡∏™‡∏≤‡∏£', '‡∏≠‡∏≠‡∏Å‡πÇ‡∏£‡∏á‡∏á‡∏≤‡∏ô', '‡∏ñ‡∏∂‡∏á‡∏™‡∏≤‡∏Ç‡∏≤', '‡∏à‡∏ö‡∏á‡∏≤‡∏ô']
    SUM_COLS = [20, 16, 20, 20, 20, 20, 20, 20, 20, 20]
    
    pdf.set_fill_color(44, 62, 80)
    pdf.set_text_color(255, 255, 255)
    pdf.set_font('Sarabun', '', 6)
    for i, h in enumerate(SUM_HEADERS):
        pdf.cell(SUM_COLS[i], 8, h, border=1, align='C', fill=True)
    pdf.ln()
    
    def draw_sum_row(label, data, is_total=False):
        if is_total: pdf.set_fill_color(255, 255, 0)
        else: pdf.set_fill_color(255, 255, 255)
        pdf.set_text_color(0, 0, 0)
        pdf.set_font('Sarabun', '', 7)
        pdf.cell(SUM_COLS[0], 8, label, border=1, align='C', fill=True)
        vals = [data['count'], data['t1'], data['t2'], data['t3'], data['t4'], data['t5'], data['t6'], data['t7'], data['t8']]
        for i, v in enumerate(vals):
            pdf.cell(SUM_COLS[i+1], 8, str(v), border=1, align='C', fill=True)
        pdf.ln()

    draw_sum_row('‡∏Å‡∏•‡∏≤‡∏á‡∏ß‡∏±‡∏ô', sum_day)
    draw_sum_row('‡∏Å‡∏•‡∏≤‡∏á‡∏Ñ‡∏∑‡∏ô', sum_night)
    draw_sum_row('‡∏£‡∏ß‡∏°', sum_total, is_total=True)

    pdf_bytes = pdf.output()
    filename = f"Summary_{date_filter if date_filter else 'All'}.pdf"
    return send_file(io.BytesIO(pdf_bytes), mimetype='application/pdf', as_attachment=True, download_name=filename)
    
@app.route('/tracking')
def customer_view():
    sheet = get_db()
    raw_jobs = get_cached_records(sheet, 'Jobs')
    all_dates = sorted(list(set([str(j['PO_Date']).strip() for j in raw_jobs])), reverse=True)
    
    date_filter = request.args.get('date_filter')
    now_thai = datetime.now() + timedelta(hours=7)
    if not date_filter: date_filter = now_thai.strftime("%Y-%m-%d")

    jobs = [j for j in raw_jobs if str(j['PO_Date']).strip() == str(date_filter).strip()]
    
    try:
        current_date_obj = datetime.strptime(date_filter, "%Y-%m-%d")
        prev_date = (current_date_obj - timedelta(days=1)).strftime("%Y-%m-%d")
        next_date = (current_date_obj + timedelta(days=1)).strftime("%Y-%m-%d")
    except ValueError:
        prev_date = date_filter
        next_date = date_filter

    jobs_by_trip_key = {}
    total_done_jobs = 0
    total_branches = len(jobs)
    
    for job in jobs:
        trip_key = (str(job['PO_Date']), str(job['Car_No']), str(job['Round']))
        if trip_key not in jobs_by_trip_key: jobs_by_trip_key[trip_key] = []
        jobs_by_trip_key[trip_key].append(job)
        if job['Status'] == 'Done': total_done_jobs += 1
            
        # [UPDATED] Logic ‡∏Ñ‡∏≥‡∏ô‡∏ß‡∏ì‡πÄ‡∏ß‡∏•‡∏≤‡πÄ‡∏Ç‡πâ‡∏≤‡∏™‡∏≤‡∏¢‡πÅ‡∏ö‡∏ö‡∏•‡∏∞‡πÄ‡∏≠‡∏µ‡∏¢‡∏î (‡∏ä‡∏±‡πà‡∏ß‡πÇ‡∏°‡∏á/‡∏ô‡∏≤‡∏ó‡∏µ)
        job['is_start_late'] = False
        job['delay_msg'] = ""
        
        t_plan_str = str(job.get('Round', '')).strip()
        t_act_str = str(job.get('T2_StartLoad', '')).strip()
        
        if t_plan_str and t_act_str:
            try:
                fmt_plan = "%H:%M" if len(t_plan_str) <= 5 else "%H:%M:%S"
                fmt_act = "%H:%M" if len(t_act_str) <= 5 else "%H:%M:%S"
                
                t_plan = datetime.strptime(t_plan_str, fmt_plan).replace(year=2000, month=1, day=1)
                t_act = datetime.strptime(t_act_str, fmt_act).replace(year=2000, month=1, day=1)
                
                if t_plan.hour >= 18 and t_act.hour < 6: 
                    t_act += timedelta(days=1)
                elif t_plan.hour < 6 and t_act.hour >= 18: 
                    t_act -= timedelta(days=1)
                    
                if t_act > t_plan: 
                    job['is_start_late'] = True
                    diff = t_act - t_plan
                    total_seconds = diff.total_seconds()
                    hours = int(total_seconds // 3600)
                    minutes = int((total_seconds % 3600) // 60)
                    job['delay_msg'] = f"‡∏•‡πà‡∏≤‡∏ä‡πâ‡∏≤ {hours} ‡∏ä‡∏°. {minutes} ‡∏ô."
            except: pass
            
    completed_trips = 0
    for trip_key, job_list in jobs_by_trip_key.items():
        if all(job['Status'] == 'Done' for job in job_list): completed_trips += 1
            
    total_trips = len(jobs_by_trip_key)
    total_running_jobs = total_branches - total_done_jobs

    def sort_key_func(job):
        car_no_str = str(job['Car_No']).strip()
        try: car_no_int = int(car_no_str)
        except ValueError: car_no_int = 99999 
        return (car_no_int)

    jobs = sorted(jobs, key=sort_key_func)
    
    return render_template('customer_view.html', 
                           jobs=jobs, all_dates=all_dates, current_date=date_filter,
                           total_trips=total_trips, completed_trips=completed_trips,
                           total_branches=total_branches, total_done_jobs=total_done_jobs,
                           total_running_jobs=total_running_jobs,
                           prev_date=prev_date, next_date=next_date)

@app.route('/driver')
def driver_select():
    # [FIXED] Use Cached Data instead of API Call
    sheet = get_db()
    cached_drivers = get_cached_records(sheet, 'Drivers')
    # Extract names from cached list (assuming 'Name' is the key)
    drivers_list_raw = [d['Name'] for d in cached_drivers if d.get('Name')]
    
    all_jobs = get_cached_records(sheet, 'Jobs')
    now_thai = datetime.now() + timedelta(hours=7)
    limit_time = now_thai + timedelta(hours=48)
    
    driver_info = {} 
    driver_sort_data = {}

    for name in drivers_list_raw:
        driver_info[name] = {
            'pending_set': set(), 
            'pending_count': 0, 
            'urgent_msg': '', 'urgent_color': '', 'urgent_time': '', 'sort_weight': 999
        }
        driver_sort_data[name] = {'dt': datetime.max, 'car': 99999}

    for job in all_jobs:
        d_name = job.get('Driver')
        if d_name not in driver_info: continue
        status = str(job.get('Status', '')).lower()
        if status != 'done' and status != 'cancel':
            trip_key = f"{job['PO_Date']}_{job['Round']}_{job['Car_No']}"
            driver_info[d_name]['pending_set'].add(trip_key)
            try:
                load_date_val = str(job.get('Load_Date', '')).strip()
                if not load_date_val: load_date_val = str(job['PO_Date']).strip()
                round_str = str(job['Round']).strip()
                job_dt_str = f"{load_date_val} {round_str}"
                try: job_dt = datetime.strptime(job_dt_str, "%Y-%m-%d %H:%M")
                except: continue
                
                try: car_n = int(str(job['Car_No']).strip())
                except: car_n = 99999

                if job_dt < driver_sort_data[d_name]['dt']:
                    driver_sort_data[d_name]['dt'] = job_dt
                    driver_sort_data[d_name]['car'] = car_n

                diff = job_dt - now_thai
                hours_diff = diff.total_seconds() / 3600
                delta_days = (job_dt.date() - now_thai.date()).days
                h = job_dt.hour
                m = job_dt.minute
                msg, color, weight = "", "", 999
                
                if hours_diff <= 0:
                    if hours_diff > -12: msg, color, weight = "‚ùó ‡πÇ‡∏´‡∏•‡∏î‡∏ï‡∏≠‡∏ô‡∏ô‡∏µ‡πâ", "bg-red-500 text-white border-red-600 animate-pulse", 1
                elif 0 < hours_diff <= 16:
                    if 6 <= h <= 12: msg, color = "‚òÄÔ∏è ‡πÇ‡∏´‡∏•‡∏î‡πÄ‡∏ä‡πâ‡∏≤‡∏ô‡∏µ‡πâ", "bg-yellow-100 text-yellow-700"
                    elif 13 <= h <= 18: msg, color = "‚õÖ ‡πÇ‡∏´‡∏•‡∏î‡∏ö‡πà‡∏≤‡∏¢‡∏ô‡∏µ‡πâ", "bg-orange-100 text-orange-700"
                    else: msg, color = "üåô ‡πÇ‡∏´‡∏•‡∏î‡∏Ñ‡∏∑‡∏ô‡∏ô‡∏µ‡πâ", "bg-indigo-100 text-indigo-700"
                    weight = 2
                elif delta_days == 1:
                    if driver_info[d_name]['sort_weight'] > 3:
                        msg, color, weight = "‚è© ‡πÄ‡∏ï‡∏£‡∏µ‡∏¢‡∏°‡∏û‡∏£‡∏∏‡πà‡∏á‡∏ô‡∏µ‡πâ", "bg-gray-100 text-gray-500", 3

                if weight < driver_info[d_name]['sort_weight']:
                    driver_info[d_name]['urgent_msg'] = msg
                    driver_info[d_name]['urgent_color'] = color
                    driver_info[d_name]['urgent_time'] = f"{h:02}:{m:02} ‡∏ô."
                    driver_info[d_name]['sort_weight'] = weight
            except: pass
    
    for name in drivers_list_raw:
        driver_info[name]['pending_count'] = len(driver_info[name]['pending_set'])

    active_drivers = []
    hidden_drivers = []
    for name in drivers_list_raw:
        earliest_dt = driver_sort_data[name]['dt']
        if earliest_dt != datetime.max and earliest_dt <= limit_time: active_drivers.append(name)
        else: hidden_drivers.append(name)
            
    def sort_key(n): return (driver_sort_data[n]['dt'], driver_sort_data[n]['car'])
    active_drivers.sort(key=sort_key)
    hidden_drivers.sort(key=sort_key)

    return render_template('driver_select.html', active_drivers=active_drivers, hidden_drivers=hidden_drivers, driver_info=driver_info)

@app.route('/driver/tasks', methods=['GET'])
def driver_tasks():
    driver_name = request.args.get('name')
    if not driver_name: return redirect(url_for('driver_select'))
        
    sheet = get_db()
    # ‡∏î‡∏∂‡∏á‡∏Ç‡πâ‡∏≠‡∏°‡∏π‡∏•‡∏ó‡∏±‡πâ‡∏á‡∏´‡∏°‡∏î (‡∏ï‡πâ‡∏≠‡∏á‡∏°‡∏±‡πà‡∏ô‡πÉ‡∏à‡∏ß‡πà‡∏≤‡πÉ‡∏ô Sheet ‡∏°‡∏µ Header: PO_Nos, Doc_Result, Weight_Result ‡πÅ‡∏•‡πâ‡∏ß)
    raw_data = get_cached_records(sheet, 'Jobs')
    
    driver_jobs_with_id = []
    for idx, job in enumerate(raw_data):
        if job['Driver'] == driver_name:
            job_copy = job.copy()
            job_copy['row_id'] = idx + 2 # ‡πÄ‡∏Å‡πá‡∏ö‡πÄ‡∏•‡∏Ç‡πÅ‡∏ñ‡∏ß‡πÄ‡∏û‡∏∑‡πà‡∏≠‡πÉ‡∏ä‡πâ Update
            driver_jobs_with_id.append(job_copy)

    # ‡∏à‡∏±‡∏î‡∏Å‡∏•‡∏∏‡πà‡∏°‡∏á‡∏≤‡∏ô‡πÄ‡∏õ‡πá‡∏ô Trip (1 Trip ‡∏°‡∏µ‡∏´‡∏•‡∏≤‡∏¢‡∏™‡∏≤‡∏Ç‡∏≤‡πÑ‡∏î‡πâ)
    trips = {}
    for job in driver_jobs_with_id:
        trip_key = (str(job['PO_Date']), str(job['Round']), str(job['Car_No']))
        if trip_key not in trips: trips[trip_key] = []
        trips[trip_key].append(job)

    final_jobs_list = []
    now_thai = datetime.now() + timedelta(hours=7)
    today_date = now_thai.date()
    
    # Logic ‡∏Å‡∏≤‡∏£‡πÄ‡∏•‡∏∑‡∏≠‡∏Å‡πÇ‡∏ä‡∏ß‡πå‡∏á‡∏≤‡∏ô: ‡πÇ‡∏ä‡∏ß‡πå‡∏á‡∏≤‡∏ô‡∏ó‡∏µ‡πà‡∏¢‡∏±‡∏á‡πÑ‡∏°‡πà‡πÄ‡∏™‡∏£‡πá‡∏à ‡∏´‡∏£‡∏∑‡∏≠ ‡∏á‡∏≤‡∏ô‡πÄ‡∏™‡∏£‡πá‡∏à‡πÅ‡∏•‡πâ‡∏ß‡∏ó‡∏µ‡πà‡πÄ‡∏õ‡πá‡∏ô‡∏Ç‡∏≠‡∏á‡∏ß‡∏±‡∏ô‡∏ô‡∏µ‡πâ/‡∏≠‡∏ô‡∏≤‡∏Ñ‡∏ï
    for key, job_list in trips.items():
        is_trip_fully_done = all(j['Status'] == 'Done' for j in job_list)
        show_this_trip = False
        
        if not is_trip_fully_done:
            show_this_trip = True
        else:
            try:
                job_date_str = job_list[0].get('Load_Date', job_list[0]['PO_Date'])
                pd = datetime.strptime(job_date_str, "%Y-%m-%d").date()
                if pd >= today_date:
                    show_this_trip = True
            except: 
                show_this_trip = True
            
        if show_this_trip:
            final_jobs_list.extend(job_list)

    # ‡πÄ‡∏£‡∏µ‡∏¢‡∏á‡∏•‡∏≥‡∏î‡∏±‡∏ö‡∏á‡∏≤‡∏ô
    def sort_key_func(job):
        return (str(job['PO_Date']), str(job.get('Load_Date', '')), str(job['Round']))
    
    my_jobs = sorted(final_jobs_list, key=sort_key_func)
    today_date_str = now_thai.strftime("%Y-%m-%d")

    # Loop ‡πÄ‡∏û‡∏∑‡πà‡∏≠‡πÄ‡∏ï‡∏£‡∏µ‡∏¢‡∏°‡∏Ç‡πâ‡∏≠‡∏°‡∏π‡∏•‡∏™‡∏≥‡∏´‡∏£‡∏±‡∏ö‡πÅ‡∏™‡∏î‡∏á‡∏ú‡∏•
    for job in my_jobs:
        # =========================================================
        # [‡∏™‡πà‡∏ß‡∏ô‡∏ó‡∏µ‡πà 1] Parsing PO Data (‡∏™‡∏≥‡∏´‡∏£‡∏±‡∏ö‡πÅ‡∏™‡∏î‡∏á‡∏ä‡πà‡∏≠‡∏á‡∏Å‡∏£‡∏≠‡∏Å‡∏Ç‡πâ‡∏≠‡∏°‡∏π‡∏•)
        # =========================================================
        po_str = str(job.get('PO_Nos', '')).strip()       # Col Z
        doc_str = str(job.get('Doc_Result', '')).strip()  # Col AA
        weight_str = str(job.get('Weight_Result', '')).strip() # Col AB
        
        job['parsed_po_details'] = []

        if po_str:
            # ‡πÅ‡∏õ‡∏•‡∏á Doc String "PO1:Doc1 | PO2:Doc2" ‡πÄ‡∏õ‡πá‡∏ô Dict
            doc_map = {}
            if doc_str:
                for p in doc_str.split('|'):
                    if ':' in p:
                        k, v = p.split(':', 1)
                        doc_map[k.strip()] = v.strip()

            # ‡πÅ‡∏õ‡∏•‡∏á Weight String "PO1:10.5 | PO2:20" ‡πÄ‡∏õ‡πá‡∏ô Dict
            weight_map = {}
            if weight_str:
                for p in weight_str.split('|'):
                    if ':' in p:
                        k, v = p.split(':', 1)
                        weight_map[k.strip()] = v.strip()

            # ‡∏™‡∏£‡πâ‡∏≤‡∏á List Object ‡πÄ‡∏û‡∏∑‡πà‡∏≠‡∏™‡πà‡∏á‡πÑ‡∏õ‡∏ß‡∏ô‡∏•‡∏π‡∏õ‡∏™‡∏£‡πâ‡∏≤‡∏á Input Field ‡πÉ‡∏ô HTML
            for po_item in po_str.split(','):
                po_name = po_item.strip()
                if po_name:
                    job['parsed_po_details'].append({
                        'name': po_name,
                        'doc': doc_map.get(po_name, ''),
                        'weight': weight_map.get(po_name, '')
                    })

        # =========================================================
        # [‡∏™‡πà‡∏ß‡∏ô‡∏ó‡∏µ‡πà 2] Smart Title & UI Decoration (‡∏Ñ‡∏≥‡∏ô‡∏ß‡∏ì‡∏™‡∏µ‡πÅ‡∏•‡∏∞‡∏™‡∏ñ‡∏≤‡∏ô‡∏∞)
        # =========================================================
        try:
            load_date_str = job.get('Load_Date', job['PO_Date'])
            round_str = str(job['Round']).strip()
            job_dt_str = f"{load_date_str} {round_str}"
            
            try: job_dt = datetime.strptime(job_dt_str, "%Y-%m-%d %H:%M")
            except: job_dt = datetime.strptime(f"{job['PO_Date']} {round_str}", "%Y-%m-%d %H:%M")
            
            diff = job_dt - now_thai
            hours_diff = diff.total_seconds() / 3600
            delta_days = (job_dt.date() - now_thai.date()).days

            th_year = job_dt.year + 543
            real_date_str = f"{job_dt.day}/{job_dt.month}/{str(th_year)[2:]}"
            h = job_dt.hour

            # ‡∏Ñ‡πà‡∏≤ Default
            job['smart_title'] = f"‡πÄ‡∏ß‡∏•‡∏≤ {round_str}"
            job['smart_detail'] = f"‡∏ß‡∏±‡∏ô‡∏ó‡∏µ‡πà {real_date_str}"
            job['ui_class'] = {'bg': 'bg-gray-50', 'text': 'text-gray-500', 'icon': 'fa-clock'}

            # 1. ‡∏á‡∏≤‡∏ô‡πÄ‡∏£‡πà‡∏á‡∏î‡πà‡∏ß‡∏ô/‡∏Ñ‡πâ‡∏≤‡∏á‡∏™‡πà‡∏á (‡∏ô‡πâ‡∏≠‡∏¢‡∏Å‡∏ß‡πà‡∏≤ 0 ‡∏ä‡∏°.)
            if hours_diff <= 0:
                if hours_diff > -12:
                    job['smart_title'] = f"‚ùó ‡πÄ‡∏Ç‡πâ‡∏≤‡πÇ‡∏´‡∏•‡∏î‡∏á‡∏≤‡∏ô‡∏ï‡∏≠‡∏ô‡∏ô‡∏µ‡πâ"
                    job['ui_class'] = {'bg': 'bg-red-50 border-red-100 ring-2 ring-red-200 animate-pulse', 'text': 'text-red-600', 'icon': 'fa-truck-ramp-box'}
                else:
                    job['smart_title'] = f"üî• ‡∏á‡∏≤‡∏ô‡∏Ñ‡πâ‡∏≤‡∏á‡∏™‡πà‡∏á"
                    job['ui_class'] = {'bg': 'bg-red-50 border-red-100', 'text': 'text-red-500', 'icon': 'fa-triangle-exclamation'}
                job['smart_detail'] = f"‡∏Å‡∏≥‡∏´‡∏ô‡∏î: {round_str} ‡∏ô. ({real_date_str})"
            
            # 2. ‡∏á‡∏≤‡∏ô‡∏†‡∏≤‡∏¢‡πÉ‡∏ô 16 ‡∏ä‡∏°. (‡πÇ‡∏´‡∏•‡∏î‡∏ß‡∏±‡∏ô‡∏ô‡∏µ‡πâ/‡∏Ñ‡∏∑‡∏ô‡∏ô‡∏µ‡πâ)
            elif 0 < hours_diff <= 16:
                if 6 <= h <= 12:    p, i, t = "‡πÄ‡∏ä‡πâ‡∏≤‡∏ô‡∏µ‡πâ", "fa-sun", "yellow"
                elif 13 <= h <= 18: p, i, t = "‡∏ö‡πà‡∏≤‡∏¢‡∏ô‡∏µ‡πâ", "fa-cloud-sun", "orange"
                else:               p, i, t = "‡∏Ñ‡∏∑‡∏ô‡∏ô‡∏µ‡πâ", "fa-moon", "indigo"
                job['smart_title'] = f"‡πÇ‡∏´‡∏•‡∏î‡∏™‡∏¥‡∏ô‡∏Ñ‡πâ‡∏≤{p}"
                job['smart_detail'] = f"‡πÄ‡∏ß‡∏•‡∏≤ {round_str} ‡∏ô. ‡∏Ç‡∏≠‡∏á‡∏ß‡∏±‡∏ô‡∏ó‡∏µ‡πà {real_date_str}"
                job['ui_class'] = {'bg': f'bg-{t}-50 border-{t}-100 ring-1 ring-{t}-50', 'text': f'text-{t}-600', 'icon': i}
            
            # 3. ‡∏á‡∏≤‡∏ô‡∏ß‡∏±‡∏ô‡∏û‡∏£‡∏∏‡πà‡∏á‡∏ô‡∏µ‡πâ
            elif delta_days == 1:
                period = "‡∏Ñ‡∏∑‡∏ô‡∏û‡∏£‡∏∏‡πà‡∏á‡∏ô‡∏µ‡πâ" if (h >= 19 or h <= 5) else "‡∏ß‡∏±‡∏ô‡∏û‡∏£‡∏∏‡πà‡∏á‡∏ô‡∏µ‡πâ"
                job['smart_title'] = f"‚è© ‡πÄ‡∏ï‡∏£‡∏µ‡∏¢‡∏°‡πÇ‡∏´‡∏•‡∏î{period}"
                job['smart_detail'] = f"‡πÄ‡∏ß‡∏•‡∏≤ {round_str} ‡∏ô. ‡∏Ç‡∏≠‡∏á‡∏ß‡∏±‡∏ô‡∏ó‡∏µ‡πà {real_date_str}"
                job['ui_class'] = {'bg': 'bg-blue-50 border-blue-100', 'text': 'text-blue-600', 'icon': 'fa-calendar-day'}
            
            # 4. ‡∏á‡∏≤‡∏ô‡∏•‡πà‡∏ß‡∏á‡∏´‡∏ô‡πâ‡∏≤
            else:
                job['smart_title'] = f"üìÖ ‡∏á‡∏≤‡∏ô‡∏•‡πà‡∏ß‡∏á‡∏´‡∏ô‡πâ‡∏≤"
                job['smart_detail'] = f"‡∏ß‡∏±‡∏ô‡∏ó‡∏µ‡πà {real_date_str} ‡πÄ‡∏ß‡∏•‡∏≤ {round_str} ‡∏ô."
                job['ui_class'] = {'bg': 'bg-gray-50 border-gray-100', 'text': 'text-gray-500', 'icon': 'fa-calendar-days'}
            
            # PO Label
            po_d = datetime.strptime(job['PO_Date'], "%Y-%m-%d")
            po_th = f"{po_d.day}/{po_d.month}/{str(po_d.year+543)[2:]}"
            job['po_label'] = f"(‡πÄ‡∏≠‡∏Å‡∏™‡∏≤‡∏£ PO ‡∏ß‡∏±‡∏ô‡∏ó‡∏µ‡πà {po_th})"
            
        except Exception as e: 
            pass

    return render_template('driver_tasks.html', name=driver_name, jobs=my_jobs, today_date=today_date_str)

@app.route('/update_status', methods=['POST'])
def update_status():
    row_id_target = int(request.form['row_id'])
    step = request.form['step']
    driver_name = request.form['driver_name']
    lat = request.form.get('lat', '')
    long = request.form.get('long', '')
    mode = request.form.get('mode', 'update')
    
    location_str = f"{lat},{long}" if lat and long else ""
    current_time = (datetime.now() + timedelta(hours=7)).strftime("%H:%M")
    
    sheet = get_db()
    ws = sheet.worksheet('Jobs')
    
    # ‡∏õ‡∏£‡∏±‡∏ö Col Index ‡∏Ç‡∏¢‡∏±‡∏ö‡πÑ‡∏õ‡∏ó‡∏≤‡∏á‡∏Ç‡∏ß‡∏≤ 1 ‡∏ä‡πà‡∏≠‡∏á (‡πÄ‡∏î‡∏¥‡∏° 8 -> 9 ‡πÄ‡∏û‡∏£‡∏≤‡∏∞‡πÅ‡∏ó‡∏£‡∏Å Weight ‡∏ó‡∏µ‡πà 8)
    time_col_map = {'1': 9, '2': 10, '3': 11, '4': 12, '5': 13, '6': 14, '7': 15, '8': 16}
    # Location ‡∏Å‡πá‡∏Ç‡∏¢‡∏±‡∏ö‡∏ï‡∏≤‡∏° (‡πÄ‡∏î‡∏¥‡∏° 17 -> 18)
    loc_col_map = {'1': 18, '2': 19, '3': 20, '4': 21, '5': 22, '6': 23, '7': 24, '8': 25}
    
    time_col = time_col_map.get(step)
    loc_col = loc_col_map.get(step)
    updates = []

    val_to_save = current_time if mode == 'update' else ""
    loc_to_save = location_str if mode == 'update' else ""

    target_row_data = ws.row_values(row_id_target)

    if step in ['1', '2', '3', '4', '5', '6']:
        if len(target_row_data) < 4: return redirect(url_for('driver_tasks', name=driver_name))
        target_po = target_row_data[0] 
        target_round = target_row_data[2]
        target_car = target_row_data[3]
        
        all_values = ws.get_all_values()
        for i, row in enumerate(all_values[1:]): 
            current_row_id = i + 2 
            if (len(row) > 3 and row[0] == target_po and row[2] == target_round and row[3] == target_car):      
                cell_coord_time = gspread.utils.rowcol_to_a1(current_row_id, time_col)
                updates.append({'range': cell_coord_time, 'values': [[val_to_save]]})
                if location_str or mode == 'cancel':
                    cell_coord_loc = gspread.utils.rowcol_to_a1(current_row_id, loc_col)
                    updates.append({'range': cell_coord_loc, 'values': [[loc_to_save]]})
        if updates: ws.batch_update(updates)

    elif step in ['7', '8']:
        cell_coord_time = gspread.utils.rowcol_to_a1(row_id_target, time_col)
        updates.append({'range': cell_coord_time, 'values': [[val_to_save]]})
        if location_str or mode == 'cancel':
            cell_coord_loc = gspread.utils.rowcol_to_a1(row_id_target, loc_col)
            updates.append({'range': cell_coord_loc, 'values': [[loc_to_save]]})
        if updates: ws.batch_update(updates)

    if step == '8': 
        # Status column ‡∏Ç‡∏¢‡∏±‡∏ö‡∏à‡∏≤‡∏Å 16 -> 17
        status_val = "Done" if mode == 'update' else ""
        ws.update_cell(row_id_target, 17, status_val)
    
    invalidate_cache('Jobs')

    # =========================================================================
    # [NEW LOGIC START] Notification Triggers
    # =========================================================================
    if mode == 'update' and len(target_row_data) > 5:
        # ‡πÄ‡∏ï‡∏£‡∏µ‡∏¢‡∏°‡∏Ç‡πâ‡∏≠‡∏°‡∏π‡∏•‡∏™‡∏≥‡∏´‡∏£‡∏±‡∏ö‡∏™‡πà‡∏á‡πÅ‡∏à‡πâ‡∏á‡πÄ‡∏ï‡∏∑‡∏≠‡∏ô
        # ‡∏ï‡∏£‡∏ß‡∏à‡∏™‡∏≠‡∏ö index ‡πÉ‡∏´‡πâ‡πÅ‡∏ô‡πà‡πÉ‡∏à‡∏ß‡πà‡∏≤‡∏ï‡∏£‡∏á‡∏Å‡∏±‡∏ö Sheet (Driver=Col E (idx 4), Plate=Col F (idx 5))
        job_info_for_notify = {
            'PO_Date': target_row_data[0],
            'Round': target_row_data[2],
            'Car_No': target_row_data[3],
            'Driver': target_row_data[4],
            'Plate': target_row_data[5]
        }

        # 1. ‡πÅ‡∏à‡πâ‡∏á‡πÄ‡∏ï‡∏∑‡∏≠‡∏ô‡∏£‡∏≤‡∏¢‡∏Ñ‡∏±‡∏ô (‡πÄ‡∏Ç‡πâ‡∏≤ Step 1 / ‡∏≠‡∏≠‡∏Å Step 6)
        if step == '1' or step == '6':
            notify_individual_movement(sheet, job_info_for_notify, step)
        # [NEW] ‡πÅ‡∏à‡πâ‡∏á‡πÄ‡∏ï‡∏∑‡∏≠‡∏ô‡∏£‡∏≤‡∏¢‡∏Ñ‡∏±‡∏ô (‡∏à‡∏ö‡∏á‡∏≤‡∏ô‡∏Ñ‡∏£‡∏ö‡∏ó‡∏∏‡∏Å‡∏™‡∏≤‡∏Ç‡∏≤ Step 8)
        if step == '8':
            notify_car_completion(sheet, job_info_for_notify)
        # 2. ‡∏ï‡∏£‡∏ß‡∏à‡∏™‡∏≠‡∏ö‡∏Å‡∏•‡∏∏‡πà‡∏° (‡πÄ‡∏Ç‡πâ‡∏≤‡∏Ñ‡∏£‡∏ö / ‡∏≠‡∏≠‡∏Å‡∏Ñ‡∏£‡∏ö / ‡∏à‡∏ö‡∏Ñ‡∏£‡∏ö)
        # ‡πÄ‡∏ä‡πá‡∏Ñ‡∏ó‡∏∏‡∏Å‡∏Ñ‡∏£‡∏±‡πâ‡∏á‡∏ó‡∏µ‡πà‡∏°‡∏µ‡∏Å‡∏≤‡∏£ update Step 1, 6 ‡∏´‡∏£‡∏∑‡∏≠ 8
        if step in ['1', '6', '8']:
            check_group_completion(sheet, target_row_data[0], target_row_data[2], step)
            
        # 3. ‡πÄ‡∏ä‡πá‡∏Ñ Late (‡∏ù‡∏≤‡∏Å‡πÄ‡∏ä‡πá‡∏Ñ‡∏ó‡∏∏‡∏Å‡∏Ñ‡∏£‡∏±‡πâ‡∏á‡∏ó‡∏µ‡πà‡∏°‡∏µ‡∏Å‡∏≤‡∏£‡∏Å‡∏î Update)
        check_late_and_notify(sheet)
    # =========================================================================
    # [NEW LOGIC END]
    # =========================================================================
        
    return redirect(url_for('driver_tasks', name=driver_name))

@app.route('/update_driver', methods=['POST'])
def update_driver():
    if 'user' not in session: return json.dumps({'status': 'error', 'message': 'Unauthorized'}), 401
    try:
        data = request.json
        target_po = data.get('po_date')
        target_round = data.get('round_time')
        target_car = str(data.get('car_no'))
        new_driver = data.get('new_driver')
        new_plate = data.get('new_plate')

        sheet = get_db()
        ws = sheet.worksheet('Jobs')
        all_values = ws.get_all_values()
        updates = []
        
        for i, row in enumerate(all_values):
            if i == 0: continue 
            if len(row) < 6: continue
            if (row[0] == target_po and str(row[2]) == str(target_round) and str(row[3]) == target_car):
                row_num = i + 1
                updates.append({'range': f'E{row_num}', 'values': [[new_driver]]})
                updates.append({'range': f'F{row_num}', 'values': [[new_plate]]})

        if updates:
            ws.batch_update(updates)
            invalidate_cache('Jobs')
            return json.dumps({'status': 'success', 'count': len(updates)/2})
        else:
            return json.dumps({'status': 'error', 'message': '‡πÑ‡∏°‡πà‡∏û‡∏ö‡∏£‡∏≤‡∏¢‡∏Å‡∏≤‡∏£‡∏á‡∏≤‡∏ô‡∏ó‡∏µ‡πà‡∏ï‡∏£‡∏á‡∏Å‡∏±‡∏ô'})
    except Exception as e:
        print(f"Error updating driver: {e}")
        return json.dumps({'status': 'error', 'message': str(e)})
        
@app.route('/save_po_detail', methods=['POST'])
def save_po_detail():
    try:
        data = request.json
        row_id = int(data.get('row_id'))
        po_name = data.get('po_name')
        val_type = data.get('type') # 'doc' ‡∏´‡∏£‡∏∑‡∏≠ 'weight'
        value = data.get('value')
        
        sheet = get_db()
        ws = sheet.worksheet('Jobs')
        
        # ‡∏≠‡πà‡∏≤‡∏ô‡∏Ç‡πâ‡∏≠‡∏°‡∏π‡∏•‡πÄ‡∏î‡∏¥‡∏°‡∏°‡∏≤‡∏Å‡πà‡∏≠‡∏ô
        # Col 27 (AA) = Doc, Col 28 (AB) = Weight
        target_col = 27 if val_type == 'doc' else 28
        current_val = ws.cell(row_id, target_col).value or ""
        
        # ‡πÅ‡∏õ‡∏•‡∏á‡πÄ‡∏õ‡πá‡∏ô Map
        val_map = {}
        if current_val:
            parts = current_val.split('|')
            for p in parts:
                if ':' in p:
                    k, v = p.split(':', 1)
                    val_map[k.strip()] = v.strip()
        
        # ‡∏≠‡∏±‡∏û‡πÄ‡∏î‡∏ó‡∏Ñ‡πà‡∏≤‡πÉ‡∏´‡∏°‡πà
        val_map[po_name] = value
        
        # ‡πÅ‡∏õ‡∏•‡∏á‡∏Å‡∏•‡∏±‡∏ö‡πÄ‡∏õ‡πá‡∏ô String "PO:Val | PO:Val"
        new_str_parts = []
        for k, v in val_map.items():
            if v: # ‡πÄ‡∏Å‡πá‡∏ö‡πÄ‡∏â‡∏û‡∏≤‡∏∞‡∏ó‡∏µ‡πà‡∏°‡∏µ‡∏Ñ‡πà‡∏≤
                new_str_parts.append(f"{k}:{v}")
        
        new_str = " | ".join(new_str_parts)
        
        # ‡∏ö‡∏±‡∏ô‡∏ó‡∏∂‡∏Å
        ws.update_cell(row_id, target_col, new_str)
        invalidate_cache('Jobs')
        
        return json.dumps({'status': 'success', 'value': value})
    except Exception as e:
        return json.dumps({'status': 'error', 'message': str(e)})
        

# ==========================================
# [UPDATED] Monthly Calendar with Midnight Crossover
# ==========================================
# ‡∏Ñ‡πâ‡∏ô‡∏´‡∏≤ route '/calendar' ‡πÅ‡∏•‡∏∞‡πÅ‡∏ó‡∏ô‡∏ó‡∏µ‡πà‡∏ü‡∏±‡∏á‡∏Å‡πå‡∏ä‡∏±‡∏ô monthly_calendar ‡∏î‡πâ‡∏ß‡∏¢‡πÇ‡∏Ñ‡πâ‡∏î‡∏ô‡∏µ‡πâ‡∏Ñ‡∏£‡∏±‡∏ö

@app.route('/calendar')
def monthly_calendar():
    # if 'user' not in session: return redirect(url_for('manager_login'))
    
    now = datetime.now() + timedelta(hours=7)
    try:
        year = int(request.args.get('year', now.year))
        month = int(request.args.get('month', now.month))
    except:
        year, month = now.year, now.month

    cal = calendar.Calendar(firstweekday=6)
    month_days = cal.monthdayscalendar(year, month)
    
    thai_months = [
        "", "‡∏°‡∏Å‡∏£‡∏≤‡∏Ñ‡∏°", "‡∏Å‡∏∏‡∏°‡∏†‡∏≤‡∏û‡∏±‡∏ô‡∏ò‡πå", "‡∏°‡∏µ‡∏ô‡∏≤‡∏Ñ‡∏°", "‡πÄ‡∏°‡∏©‡∏≤‡∏¢‡∏ô", "‡∏û‡∏§‡∏©‡∏†‡∏≤‡∏Ñ‡∏°", "‡∏°‡∏¥‡∏ñ‡∏∏‡∏ô‡∏≤‡∏¢‡∏ô",
        "‡∏Å‡∏£‡∏Å‡∏é‡∏≤‡∏Ñ‡∏°", "‡∏™‡∏¥‡∏á‡∏´‡∏≤‡∏Ñ‡∏°", "‡∏Å‡∏±‡∏ô‡∏¢‡∏≤‡∏¢‡∏ô", "‡∏ï‡∏∏‡∏•‡∏≤‡∏Ñ‡∏°", "‡∏û‡∏§‡∏®‡∏à‡∏¥‡∏Å‡∏≤‡∏¢‡∏ô", "‡∏ò‡∏±‡∏ô‡∏ß‡∏≤‡∏Ñ‡∏°"
    ]
    month_name = thai_months[month]

    sheet = get_db()
    raw_jobs = get_cached_records(sheet, 'Jobs')
    raw_drivers = get_cached_records(sheet, 'Drivers')
    all_driver_names = [d['Name'] for d in raw_drivers if d.get('Name')]

    # ‡πÄ‡∏Å‡πá‡∏ö‡∏Ç‡πâ‡∏≠‡∏°‡∏π‡∏•‡πÄ‡∏õ‡πá‡∏ô Set ‡πÄ‡∏û‡∏∑‡πà‡∏≠‡∏Å‡∏±‡∏ô‡∏ã‡πâ‡∏≥ (Trip ID)
    # ‡πÇ‡∏Ñ‡∏£‡∏á‡∏™‡∏£‡πâ‡∏≤‡∏á: calendar_data[day] = { 'day': { 'DriverName': {set_of_trip_ids} }, ... }
    calendar_data = {}

    for job in raw_jobs:
        if str(job.get('Status', '')).lower() == 'cancel': continue
        
        date_str = job.get('Load_Date', '').strip()
        if not date_str: date_str = str(job['PO_Date']).strip()
        time_str = str(job['Round']).strip()
        
        try:
            if not time_str: time_str = "12:00"
            fmt_time = "%H:%M" if len(time_str) <= 5 else "%H:%M:%S"
            job_dt = datetime.strptime(f"{date_str} {time_str}", f"%Y-%m-%d {fmt_time}")
            
            # Logic: Midnight Crossover
            if 0 <= job_dt.hour < 6:
                job_dt = job_dt - timedelta(days=1)
            
            if job_dt.year != year or job_dt.month != month:
                continue
                
            day_key = job_dt.day
            
            if day_key not in calendar_data:
                calendar_data[day_key] = {'day_drivers': {}, 'night_drivers': {}}
            
            is_day = True
            h = int(time_str.split(':')[0])
            if h < 6 or h >= 19: is_day = False
            
            driver_name = job['Driver']
            
            # ‡∏™‡∏£‡πâ‡∏≤‡∏á Unique ID ‡∏™‡∏≥‡∏´‡∏£‡∏±‡∏ö‡πÄ‡∏ó‡∏µ‡πà‡∏¢‡∏ß‡∏ß‡∏¥‡πà‡∏á‡∏ô‡∏µ‡πâ (‡πÄ‡∏ä‡πà‡∏ô "11:00_14")
            # ‡∏ñ‡πâ‡∏≤‡πÄ‡∏õ‡πá‡∏ô‡∏£‡∏≠‡∏ö‡πÄ‡∏î‡∏µ‡∏¢‡∏ß‡∏Å‡∏±‡∏ô ‡∏Ñ‡∏±‡∏ô‡πÄ‡∏î‡∏µ‡∏¢‡∏ß‡∏Å‡∏±‡∏ô ‡πÅ‡∏°‡πâ‡∏à‡∏∞‡∏°‡∏µ‡∏´‡∏•‡∏≤‡∏¢‡∏™‡∏≤‡∏Ç‡∏≤ ID ‡∏ô‡∏µ‡πâ‡∏à‡∏∞‡πÄ‡∏´‡∏°‡∏∑‡∏≠‡∏ô‡πÄ‡∏î‡∏¥‡∏°
            trip_id = f"{time_str}_{job.get('Car_No', '')}"

            if is_day:
                if driver_name not in calendar_data[day_key]['day_drivers']:
                    calendar_data[day_key]['day_drivers'][driver_name] = set()
                calendar_data[day_key]['day_drivers'][driver_name].add(trip_id)
            else:
                if driver_name not in calendar_data[day_key]['night_drivers']:
                    calendar_data[day_key]['night_drivers'][driver_name] = set()
                calendar_data[day_key]['night_drivers'][driver_name].add(trip_id)
                
        except Exception as e: 
            continue

    # ‡∏à‡∏±‡∏î‡∏£‡∏π‡∏õ‡πÅ‡∏ö‡∏ö‡∏Ç‡πâ‡∏≠‡∏°‡∏π‡∏•‡∏™‡∏∏‡∏î‡∏ó‡πâ‡∏≤‡∏¢
    final_data = {}
    for d in range(1, 32):
        if d in calendar_data:
            day_active_dict = calendar_data[d]['day_drivers']
            night_active_dict = calendar_data[d]['night_drivers']
            
            # ‡πÅ‡∏õ‡∏•‡∏á Set ‡πÉ‡∏´‡πâ‡πÄ‡∏õ‡πá‡∏ô‡∏à‡∏≥‡∏ô‡∏ß‡∏ô‡∏ô‡∏±‡∏ö (len)
            # ‡∏ñ‡πâ‡∏≤‡∏°‡∏µ 2 ‡∏™‡∏≤‡∏Ç‡∏≤ ‡πÅ‡∏ï‡πà trip_id ‡πÄ‡∏î‡∏µ‡∏¢‡∏ß‡∏Å‡∏±‡∏ô set ‡∏à‡∏∞‡∏°‡∏µ‡∏Ç‡∏ô‡∏≤‡∏î‡πÅ‡∏Ñ‡πà 1 -> ‡∏ô‡∏±‡∏ö‡πÄ‡∏õ‡πá‡∏ô 1 ‡∏£‡∏≠‡∏ö (‡∏ñ‡∏π‡∏Å‡∏ï‡πâ‡∏≠‡∏á)
            # ‡∏ñ‡πâ‡∏≤‡∏°‡∏µ 2 ‡∏£‡∏≠‡∏ö (11:00 ‡πÅ‡∏•‡∏∞ 18:30) trip_id ‡∏à‡∏∞‡∏ï‡πà‡∏≤‡∏á‡∏Å‡∏±‡∏ô -> ‡∏ô‡∏±‡∏ö‡πÄ‡∏õ‡πá‡∏ô 2 ‡∏£‡∏≠‡∏ö (‡∏ñ‡∏π‡∏Å‡∏ï‡πâ‡∏≠‡∏á)
            
            day_active = sorted(
                [{'name': k, 'count': len(v)} for k, v in day_active_dict.items()], 
                key=lambda x: x['name']
            )
            
            night_active = sorted(
                [{'name': k, 'count': len(v)} for k, v in night_active_dict.items()], 
                key=lambda x: x['name']
            )
            
            active_day_names = set(day_active_dict.keys())
            active_night_names = set(night_active_dict.keys())
            
            day_standby = sorted(list(set(all_driver_names) - active_day_names))
            night_standby = sorted(list(set(all_driver_names) - active_night_names))
            
            final_data[d] = {
                'day_active': day_active,
                'night_active': night_active,
                'day_standby': day_standby,
                'night_standby': night_standby,
                'day_count': len(day_active),
                'night_count': len(night_active)
            }

    prev_month = month - 1 if month > 1 else 12
    prev_year = year if month > 1 else year - 1
    next_month = month + 1 if month < 12 else 1
    next_year = year if month < 12 else year + 1

    return render_template(
        'calendar.html',
        calendar_matrix=month_days,
        data=final_data,
        year=year,
        month=month,
        month_name=month_name,
        prev_year=prev_year, prev_month=prev_month,
        next_year=next_year, next_month=next_month,
        current_day=now.day if (now.year==year and now.month==month) else None
    )

@app.route('/')
def index(): return render_template('index.html')

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('index'))

if __name__ == '__main__':
    app.run(debug=True)